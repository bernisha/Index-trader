# -*- coding: utf-8 -*-
"""
Created on Thu Mar 29 15:02:40 2018

@author: blala
"""

"""
'******************************************************************************************************************************************************************************    
'                                                      Create asset classifications (multiple levels)
                                                                 for futures trades
'******************************************************************************************************************************************************************************    
"""

def assetClassFO(Sec_type, ins_code,sec_nam,Type_Fund, cash_flows_eff):
    
        #ssf=['OMLS'+str((cash_flows_eff['fut_sufx'].values)[0]), 'OMAS'+str((cash_flows_eff['fut_sufx'].values)[0])]
        ssf=['S']
        #excp=['OMLF'+str((cash_flows_eff['fut_sufx'].values)[0]),'OMAF'+str((cash_flows_eff['fut_sufx'].values)[0])]
        excp=['F']
        ind_fut=[str((cash_flows_eff['fut_sufx'].values)[0])] # index future suffix
        
        if Sec_type == 'CASH : CALL ACC':
            return "Total cash,Settled cash,Cash on call,Total cash"
        elif Sec_type=='CASH : SAFEX AC':
            return "Total cash,Settled cash,Futures margin,Total cash"
        elif Sec_type == "CURRENCY" and sec_nam=='VAL':
            return "Total cash,Settled cash,Val cash,Total cash"
        elif Sec_type=="PAYABLE" and sec_nam=='DIF':
            return "Total cash,Unsettled cash,Dif cash,Total cash"
        elif Sec_type=='FUTRE STCK INDX':
            return str("Futures Exposure,"+"Index Future,"+str(ins_code[0:4]+ind_fut[0])+",Futures Exposure")
    #    elif Sec_type=='FUTURE : EQUITY' and ins_code in(ssf) :
    #    elif Sec_type=='FUTURE : EQUITY' and ins_code[3:4] in(ssf):
        elif Sec_type=='FUTURE : EQUITY':
    #        return str("Futures Exposure,"+"SSF,"+str(ssf[0]))
            return str("Futures Exposure,"+"SSF,null"+",Futures Exposure")
        elif Sec_type=='EQ : ORDINARY':
            return "Equity Exposure,Equity,null,Equity Exposure"
        elif Sec_type=='EQ : RIGHTS':
            return "Equity Exposure,Equity Rights,null,Equity Exposure"
        elif Sec_type=='EQ : FOREIGN' and Type_Fund != 'M':
            return "Equity Exposure,Equity Foreign,null,Equity Exposure"
        elif ins_code[3:4] in(excp):
    #        return str("Dividend Exposure,"+"SSF Div,"+str(excp[0]))
            return str("Dividend Exposure,"+"SSF Div,null,Dividend Exposure")
        elif Sec_type=="FUND : LOC EQ":
            return str("Equity Exposure,"+"Equity Fund,"+str(ins_code)+",Equity Exposure")
        else:
            return "Other,null,null,Other"


def assetClassF(Sec_type, ins_code,sec_nam,Type_Fund, cash_flows_eff):
    
        #ssf=['OMLS'+str((cash_flows_eff['fut_sufx'].values)[0]), 'OMAS'+str((cash_flows_eff['fut_sufx'].values)[0])]
        ssf=['S']
        #excp=['OMLF'+str((cash_flows_eff['fut_sufx'].values)[0]),'OMAF'+str((cash_flows_eff['fut_sufx'].values)[0])]
        excp=['F']
        ind_fut=[str((cash_flows_eff['fut_sufx'].values)[0])] # index future suffix
        
        if Sec_type == 'CASH : CALL ACCOUNT':
            return "Total cash,Settled cash,Cash on call,Total cash"
        elif Sec_type=='CASH : SAFEX MARGIN ACCOUNT':
            return "Total cash,Settled cash,Futures margin,Total cash"
        elif Sec_type == "CURRENCY" and sec_nam=='VAL':
            return "Total cash,Settled cash,Val cash,Total cash"
        elif Sec_type=="PAYABLE" and sec_nam=='DIF':
            return "Total cash,Unsettled cash,Dif cash,Total cash"
        elif Sec_type=='FUTURE : EQUITY INDEX':
            return str("Futures Exposure,"+"Index Future,"+str(ins_code[0:4]+ind_fut[0])+",Futures Exposure")
    #    elif Sec_type=='FUTURE : EQUITY' and ins_code in(ssf) :
    #    elif Sec_type=='FUTURE : EQUITY' and ins_code[3:4] in(ssf):
        elif (Sec_type=='FUTURE : EQUITY'):
    #        return str("Futures Exposure,"+"SSF,"+str(ssf[0]))
            return str("Futures Exposure,"+"SSF,null"+",Futures Exposure")
        elif Sec_type=='EQ : ORDINARY SHARE':
            return "Equity Exposure,Equity,null,Equity Exposure"
        elif Sec_type=='EQ : STANDARD RIGHTS ISSUE':
            return "Equity Exposure,Equity Rights,null,Equity Exposure"
        elif Sec_type=='EQ : FOREIGN EQUITY' and Type_Fund != 'M':
            return "Equity Exposure,Equity Foreign,null,Equity Exposure"
        elif ins_code[3:4] in(excp):
    #        return str("Dividend Exposure,"+"SSF Div,"+str(excp[0]))
            return str("Dividend Exposure,"+"SSF Div,null,Dividend Exposure")
        elif Sec_type=="FUND : LOCAL EQUITY":
            return str("Equity Exposure,"+"Equity Fund,"+str(ins_code)+",Equity Exposure")
        else:
            return "Other,null,null,Other"


            
            
def res_indF(dat,des,ind=['Trade_date','Port_code','AssetType1','AssetType2','AssetType3','AssetType4','Quantity','EffExposure','MarketValue','FundValue','Close_price']):
    dat=dat.reset_index()
    dat['AssetType1']=des
    dat['AssetType2']='null'
    dat['AssetType3']='null'
    dat['AssetType4']='null'
    dat=dat[ind]
    return dat


def assetClassBO(Sec_type, ins_code,sec_nam, Type_Fund,cash_flows_eff):

    #ssf=['OMLS'+str((cash_flows_eff['fut_sufx'].values)[0]), 'OMAS'+str((cash_flows_eff['fut_sufx'].values)[0])]
    ssf=['S']
    #excp=['OMLF'+str((cash_flows_eff['fut_sufx'].values)[0]),'OMAF'+str((cash_flows_eff['fut_sufx'].values)[0])]
    excp=['F']
    ind_fut=[str((cash_flows_eff['fut_sufx'].values)[0])] # index future suffix
    
    if Sec_type == 'CASH : CALL ACC':
        return "A. Total cash,Settled cash,Cash on call,Total cash,C. CALL"
    elif Sec_type=='CASH : SAFEX AC':
        return "A. Total cash,Settled cash,Futures margin,Total cash,D. SAFEX"
    elif Sec_type == "CURRENCY" and sec_nam=='VAL':
        return "A. Total cash,Settled cash,Val cash,Total cash,A. VAL"
    elif Sec_type=="PAYABLE" and sec_nam=='DIF':
        return "A. Total cash,Unsettled cash,Dif cash,Total cash,B. DIF"
    elif Sec_type=='FUTRE STCK INDX':
        return str("B. Futures Exposure,"+"Index Future,"+str(ins_code[0:4]+ind_fut[0])+",Futures Exposure,A. INDEX FUTURES")


#    elif Sec_type=='FUTURE : EQUITY' and ins_code in(ssf) :
#    elif Sec_type=='FUTURE : EQUITY' and ins_code[3:4] in(ssf):
    elif Sec_type=='FUTURE : EQUITY':
#        return str("Futures Exposure,"+"SSF,"+str(ssf[0]))
        return str("B. Futures Exposure,"+"SSF,null"+",Futures Exposure"+",B. SSF")
    elif Sec_type=='EQ : ORDINARY':
        return "Equity Exposure,Equity,null,Equity Exposure,EQUITY"
    elif Sec_type=='EQ : RIGHTS':
        return "Equity Exposure,Equity Rights,null,Equity Exposure,EQUITY"
    elif Sec_type=='EQ : FOREIGN' and Type_Fund !='M':
        return "Equity Exposure,Equity Foreign,null,Equity Exposure,EQUITY"
    elif ins_code[3:4] in(excp):
#        return str("Dividend Exposure,"+"SSF Div,"+str(excp[0]))
        return str("Dividend Exposure,"+"SSF Div,null,Dividend Exposure,SSF DIV")
    elif Sec_type=="FUND : LOC EQ":
        return str("Equity Exposure,"+"Equity Fund,"+str(ins_code)+",Equity Exposure,EQUITY")
    else:
        return "Other,null,null,Other,OTHER"
        
def assetClassB(Sec_type, ins_code,sec_nam, Type_Fund,cash_flows_eff):

    #ssf=['OMLS'+str((cash_flows_eff['fut_sufx'].values)[0]), 'OMAS'+str((cash_flows_eff['fut_sufx'].values)[0])]
    ssf=['S']
    #excp=['OMLF'+str((cash_flows_eff['fut_sufx'].values)[0]),'OMAF'+str((cash_flows_eff['fut_sufx'].values)[0])]
    excp=['F']
    ind_fut=[str((cash_flows_eff['fut_sufx'].values)[0])] # index future suffix
    
    if Sec_type == 'CASH : CALL ACCOUNT':
        return "A. Total cash,Settled cash,Cash on call,Total cash,C. CALL"
    elif Sec_type=='CASH : SAFEX MARGIN ACCOUNT':
        return "A. Total cash,Settled cash,Futures margin,Total cash,D. SAFEX"
    elif Sec_type == "CURRENCY" and sec_nam=='VAL':
        return "A. Total cash,Settled cash,Val cash,Total cash,A. VAL"
    elif Sec_type=="PAYABLE" and sec_nam=='DIF':
        return "A. Total cash,Unsettled cash,Dif cash,Total cash,B. DIF"
    elif Sec_type=='FUTURE : EQUITY INDEX':
        return str("B. Futures Exposure,"+"Index Future,"+str(ins_code[0:4]+ind_fut[0])+",Futures Exposure,A. INDEX FUTURES")


#    elif Sec_type=='FUTURE : EQUITY' and ins_code in(ssf) :
 #    elif Sec_type=='FUTURE : EQUITY' and ins_code[3:4] in(ssf):
    elif (Sec_type=='FUTURE : EQUITY'):
#        return str("Futures Exposure,"+"SSF,"+str(ssf[0]))
        return str("B. Futures Exposure,"+"SSF,null"+",Futures Exposure"+",B. SSF")
    elif Sec_type=='EQ : ORDINARY SHARE':
        return "Equity Exposure,Equity,null,Equity Exposure,EQUITY"
    elif Sec_type=='EQ : STANDARD RIGHTS ISSUE':
        return "Equity Exposure,Equity Rights,null,Equity Exposure,EQUITY"
    elif Sec_type=='EQ : FOREIGN EQUITY' and Type_Fund !='M':
        return "Equity Exposure,Equity Foreign,null,Equity Exposure,EQUITY"
    elif ins_code[3:4] in(excp):
#        return str("Dividend Exposure,"+"SSF Div,"+str(excp[0]))
        return str("Dividend Exposure,"+"SSF Div,null,Dividend Exposure,SSF DIV")
    elif Sec_type=="FUND : LOCAL EQUITY":
        return str("Equity Exposure,"+"Equity Fund,"+str(ins_code)+",Equity Exposure,EQUITY")
    else:
        return "Other,null,null,Other,OTHER"
            
    
def res_indB(dat,des,ind=['Trade_date','Port_code','AssetType1','AssetType2','AssetType3','AssetType4','AssetType5','Quantity','EffExposure','MarketValue','FundValue','Close_price']):
    dat=dat.reset_index()
    dat['AssetType1']=des
    dat['AssetType2']='null'
    dat['AssetType3']='null'
    dat['AssetType4']='null'
    dat['AssetType5']='null'
    dat=dat[ind]
    return dat

"""
'******************************************************************************************************************************************************************************    
'                                                      FX to create trade & portfolio level statistics - futures
'******************************************************************************************************************************************************************************    
"""

    


def fx_dtaF(dfprt_x,  startDate):
    import pandas as pd
    import numpy as np
    from write_excel import res_indF as res_ind
    
    dfprt_1=dfprt_x.groupby(['Trade_date','Port_code','AssetType1','AssetType2','AssetType3']).agg({'EffExposure':'sum','MarketValue':'sum','Quantity':'sum','Close_price':'max'})
    dfprt_1=dfprt_1.reset_index()
    dfprt_2= (dfprt_1.groupby(['Trade_date','Port_code']).agg({'MarketValue':'sum'})).reset_index()
    dfprt_1=pd.merge( dfprt_1,dfprt_2, on=['Trade_date','Port_code'])
    dfprt_1.rename(columns={'MarketValue_x':'MarketValue', 'MarketValue_y':'FundValue'}, inplace=True)
    dfprt_1=dfprt_1[['Trade_date','Port_code','AssetType1','AssetType2','AssetType3','MarketValue','EffExposure','Quantity','FundValue','Close_price']]
    dfprt_1=dfprt_1.groupby(['Trade_date','Port_code','AssetType1','AssetType2','AssetType3']).agg({'EffExposure':'sum','MarketValue':'sum','FundValue':'max','Quantity':'max','Close_price':'max'})
    
    req_sum={'EffExposure':'sum','MarketValue':'sum','FundValue':'max','Quantity':'max','Close_price':'max'}
    total_cash= (dfprt_1[(dfprt_1.index.get_level_values('AssetType1').isin(['Total cash']))]).reset_index().groupby(['Trade_date','Port_code']).agg(req_sum)
    
    effective_cash=((total_cash-(dfprt_1[(dfprt_1.index.get_level_values('AssetType1').isin(['Futures Exposure']))]).reset_index().groupby(['Trade_date','Port_code']).agg(req_sum)).fillna(0))
    effective_cash['MarketValue']=0
    effective_cash['FundValue']=total_cash[['FundValue']].values
    effective_cash['EffExposure']=np.where(effective_cash[['EffExposure']].values==0,total_cash[['EffExposure']].values, effective_cash[['EffExposure']].values)
    
    
    cash_dat=res_ind(effective_cash,'Effective cash').reset_index()
    cash_dat['Trade_date']=startDate
    cash_dat=(cash_dat[['Trade_date', 'Port_code','AssetType1','AssetType2','AssetType3', 'Quantity','EffExposure','MarketValue','FundValue','Close_price']])
    new_dat=((pd.concat([dfprt_1.reset_index(),cash_dat],axis=0,sort=True).reset_index().drop('index',axis=1)).sort_values(['Port_code','AssetType1','AssetType2','AssetType3'])).set_index(['Trade_date','Port_code','AssetType1','AssetType2','AssetType3'])
    new_dat['EffWgt']=new_dat[['EffExposure']].values/new_dat[['FundValue']].values
    new_dat['MktWgt']=new_dat[['MarketValue']].values/new_dat[['FundValue']].values
    n_1 = new_dat.reset_index()
    n_1=n_1.groupby(['Port_code','AssetType1']).agg({'EffExposure':'sum','EffWgt':'sum'})
    n_1=n_1[~(n_1.index.get_level_values('AssetType1').isin(['Dividend Exposure']))]
    n_2=n_1.reset_index()
    fnd_value=(total_cash[['FundValue']].reset_index().set_index('Port_code')[['FundValue']]).reset_index()
    fnd_value['AssetType1']='Fund Value'
    fnd_value['EffWgt']=1
    fnd_value.columns= ['Port_code','EffExposure','AssetType1','EffWgt']  
    fnd_value=fnd_value[n_2.columns]
    n_3=n_2.append(fnd_value)
    n_3=n_3.reset_index().pivot(index='Port_code', columns='AssetType1',values='EffExposure')
    n_4=n_2.reset_index().pivot(index='Port_code', columns='AssetType1',values='EffWgt')
    
    n_3.columns=[sym.replace(" ", "")+'_R'  for sym in n_3.columns]
    n_4.columns=[sym.replace(" ", "")+'_p'  for sym in n_4.columns]
    
    n_comb=n_3.merge(n_4, left_index=True, right_index=True)   
    n_comb[['FuturesExposure_R']]=(n_comb[['FuturesExposure_R']]).fillna(0)   
    n_comb[['FuturesExposure_p']]=(n_comb[['FuturesExposure_p']]).fillna(0)   
    lst = [new_dat, n_comb]
    return lst

def fx_dtaB(dfprt_x, startDate):
    import pandas as pd
    import numpy as np
    from write_excel import res_indB as res_ind
 
    
    dfprt_1=dfprt_x.groupby(['Trade_date','Port_code','AssetType1','AssetType5','AssetType3']).agg({'EffExposure':'sum','MarketValue':'sum','Quantity':'sum','Close_price':'max'})
    dfprt_1=dfprt_1.reset_index()
    dfprt_2= (dfprt_1.groupby(['Trade_date','Port_code']).agg({'MarketValue':'sum'})).reset_index()
    dfprt_1=pd.merge( dfprt_1,dfprt_2, on=['Trade_date','Port_code'])
    dfprt_1.rename(columns={'MarketValue_x':'MarketValue', 'MarketValue_y':'FundValue'}, inplace=True)
    dfprt_1=dfprt_1[['Trade_date','Port_code','AssetType1','AssetType5','AssetType3','MarketValue','EffExposure','Quantity','FundValue','Close_price']]
    dfprt_1=dfprt_1.groupby(['Trade_date','Port_code','AssetType1','AssetType5','AssetType3']).agg({'EffExposure':'sum','MarketValue':'sum','FundValue':'max','Quantity':'max','Close_price':'max'})
    
    req_sum={'EffExposure':'sum','MarketValue':'sum','FundValue':'max','Quantity':'max','Close_price':'max'}
    total_cash= (dfprt_1[(dfprt_1.index.get_level_values('AssetType1').isin(['A. Total cash']))]).reset_index().groupby(['Trade_date','Port_code']).agg(req_sum)
    
    effective_cash=((total_cash-(dfprt_1[(dfprt_1.index.get_level_values('AssetType1').isin(['B. Futures Exposure']))]).reset_index().groupby(['Trade_date','Port_code']).agg(req_sum)).fillna(0))
    effective_cash['MarketValue']=0
    effective_cash['FundValue']=total_cash[['FundValue']].values
    effective_cash['EffExposure']=np.where(effective_cash[['EffExposure']].values==0,total_cash[['EffExposure']].values, effective_cash[['EffExposure']].values)
    
    
    cash_dat=res_ind(effective_cash,'Effective cash').reset_index()
    cash_dat['Trade_date']=startDate
    cash_dat=(cash_dat[['Trade_date', 'Port_code','AssetType1','AssetType5','AssetType3', 'Quantity','EffExposure','MarketValue','FundValue','Close_price']])
    new_dat=((pd.concat([dfprt_1.reset_index(),cash_dat],axis=0, sort=True).reset_index().drop('index',axis=1)).sort_values(['Port_code','AssetType1','AssetType5','AssetType3'])).set_index(['Trade_date','Port_code','AssetType1','AssetType5','AssetType3'])
    new_dat['EffWgt']=new_dat[['EffExposure']].values/new_dat[['FundValue']].values
    new_dat['MktWgt']=new_dat[['MarketValue']].values/new_dat[['FundValue']].values
    n_1 = new_dat.reset_index()
    n_1=n_1.groupby(['Port_code','AssetType1']).agg({'EffExposure':'sum','EffWgt':'sum'})
    n_1=n_1[~(n_1.index.get_level_values('AssetType1').isin(['Dividend Exposure']))]
    n_2=n_1.reset_index()
    fnd_value=(total_cash[['FundValue']].reset_index().set_index('Port_code')[['FundValue']]).reset_index()
    fnd_value['AssetType1']='Fund Value'
    fnd_value['EffWgt']=1
    fnd_value.columns= ['Port_code','EffExposure','AssetType1','EffWgt']  
    fnd_value=fnd_value[n_2.columns]
    n_3=n_2.append(fnd_value)
    n_3=n_3.reset_index().pivot(index='Port_code', columns='AssetType1',values='EffExposure')
    n_4=n_2.reset_index().pivot(index='Port_code', columns='AssetType1',values='EffWgt')
    
    n_3.columns=[sym.replace(" ", "")+'_R'  for sym in n_3.columns]
    n_4.columns=[sym.replace(" ", "")+'_p'  for sym in n_4.columns]
    
    n_comb=n_3.merge(n_4, left_index=True, right_index=True)   
    n_comb[['B.FuturesExposure_R']]=(n_comb[['B.FuturesExposure_R']]).fillna(0)   
    n_comb[['B.FuturesExposure_p']]=(n_comb[['B.FuturesExposure_p']]).fillna(0)   
    lst = [new_dat, n_comb]
    return lst

"""
'******************************************************************************************************************************************************************************    
'                                                      Create a check so futures trade is not below the minimum 
                '                                                       effective cash
'******************************************************************************************************************************************************************************    
"""



def chck_fut(no_fut, eff_cash, mx_eff_cash, mn_eff_cash, tgt_eff_cash, cls_price, fnd_value):
        eff_cash_pt=(-(no_fut*cls_price*10)/fnd_value)+eff_cash
        #print(str(eff_cash_pt))
        cnt=1
        
        while ((eff_cash_pt < mn_eff_cash)&(cnt<10)):
            no_fut=no_fut-1
            eff_cash_pt = (-(no_fut*cls_price*10)/fnd_value)+eff_cash
            cnt=cnt+1
           # print("Futures:"+str(no_fut)+", Eff cash"+str(eff_cash_pt))
            return no_fut
     
            break
        else:
            return no_fut

"""
'******************************************************************************************************************************************************************************    
'                                                      Create a Pandas Excel writer using XlsxWriter as the engine
'                                                       Futures report with required number of futures trades
'******************************************************************************************************************************************************************************    
"""


def excel_fx(output_folder,dic_users,n_comb_eff_1,startDate,newest,orders):
    
    import pandas as pd
    import numpy as np
    import datetime as dt
    import os
    from datetime import datetime, timedelta
    import openpyxl as px
    from openpyxl.styles import colors, Font, Border, Side ,Protection
    import time
    from write_excel import select_fund as sf 
    
    output_file = output_folder+'\\IndexFutRep_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsx'
    st_row = 7
    st_it = st_row+1
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
    hdr= ['FundValue', 'EquityExposure', 'Totalcash', 'FuturesExposure','Effectivecash', 
          'Cash Flow','Totalcash', 'Effectivecash',
          'Tgt_EffCash', 'TradeValue','No. Futures / Price', 'FutureCode','Trade','FundValue', 'EquityExposure', 'Totalcash',
          'FuturesExposure', 'Effectivecash', 'Check cash', 'TradeSignal','TradeComment','Checked by']        
    lst_fund= sf(struc=False)
    
    n_comb_eff_1=n_comb_eff_1[(n_comb_eff_1.index.get_level_values('Port_code').isin(lst_fund))]
    
    
    n_comb_dta=n_comb_eff_1[['FundValue_R_pf', 'EquityExposure_R_pf', 'Totalcash_R_pf', 'FuturesExposure_R_pf','Effectivecash_R_pf',
                            'Inflow', 'Totalcash_R', 'Effectivecash_R',
                       #     'FundValue_R', 'EquityExposure_R', 'Totalcash_R', 'FuturesExposure_R','Effectivecash_R',
                            'Tgt_EffCash1', 'FuturesTraded_R', 'No. Futures', 'AssetType3','Trade',
                            'FundValue_TR', 'EquityExposure_TR', 'Totalcash_TR','FuturesExposure_TR', 'Effectivecash_TR', 'Check cash','Trade_YN','Comment','Checked']]
    
    n_comb_dta.to_excel(writer, sheet_name='Sheet1', startrow=st_row, header=  hdr,index_label = ['Portfolio Code',' '])
    
    
    workbook  = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # colour source (hex codes): http://dmcritchie.mvps.org/excel/colors.htm
    # Convert the dataframe to an XlsxWriter Excel object.
    cell_format1 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':12})
    cell_format2 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':11})
    cell_format2_1 = workbook.add_format({'bold': False, 'font_color': 'black', 'font':11})
    cell_format2_2 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':11,  'border':1})
    cell_format2_2.set_text_wrap() 
    cell_format2_2.set_font_name('Calibri')
    cell_format3 = workbook.add_format({'bold': True, 'bg_color':'#FFFF99', 'font':11, 'locked':False })
    cell_format4 = workbook.add_format({'bold': True, 'bg_color':'#339966', 'font':11, 'locked':False })
    unlocked = workbook.add_format({'locked': False})
    locked = workbook.add_format({'locked': True})
    cell_format1_1 = workbook.add_format({'bold': False, 'font_color': 'white', 'font':12})
    cell_format5 = workbook.add_format({'bold': True, 'bg_color':'#00FFFF', 'font':12})
    
    
    worksheet.write_string('A1', 'Indexation Futures Report',cell_format1)
  
    #        worksheet.write('A2', '<portfolio owner>')
    worksheet.write('A3', 'Date', cell_format2)
    #worksheet.write('B1', fund)
    #worksheet.write('B2', Manager)
    worksheet.write('B3', datetime.strftime(datetime.today(), "%Y-%m-%d %H:%M:%S"), cell_format2_1)
    worksheet.write_string('A4', 'Prepared by',cell_format2)
    worksheet.write_string('B4',str(dic_users[os.environ.get("USERNAME").lower()][1]).upper(), cell_format2_1)
    worksheet.write_string('A6', 'Authorised by',cell_format2)
    worksheet.merge_range('B6:C6','', cell_format3)
    worksheet.write_string('D3', 'File used:',cell_format2)
    worksheet.write_string('E3', newest,cell_format2_1)
    worksheet.write_string('D4', 'Timestamp:',cell_format2)
    worksheet.write_string('E4', time.ctime(os.path.getmtime(newest)) ,cell_format2_1)
    worksheet.write_string('A5', 'Orders included',cell_format2)
    worksheet.write_string('B5', str(orders),cell_format5)

    #worksheet.write_string('A7', 'USWIMF',merge_format1)
      
    
    merge_format1 = workbook.add_format({'bold': 1,'border': 0,'align': 'center','valign': 'vcenter','fg_color': '#C6EFCE', 'font_color':'#006100'})
    merge_format2 = workbook.add_format({'bold': 1,'border': 0,'align': 'center','valign': 'vcenter','fg_color': '#CCFFFF', 'font_color':'#003366'}) # blue
    merge_format3 = workbook.add_format({'bold': 1,'border': 0,'align': 'center','valign': 'vcenter','fg_color': '#FF99CC', 'font_color':'#993366'}) # pink
    merge_format4 = workbook.add_format({'bold': 1,'fg_color': '#C6EFCE', 'font_color':'#006100','num_format': '0.000%'})
    merge_format5 = workbook.add_format({'bold': 1,'border': 0,'align': 'center','valign': 'vcenter','fg_color': '#CCCCFF', 'font_color':'#800080'}) # blue
    merge_format6 = workbook.add_format({'bold': 1,'border': 0,'align': 'center','valign': 'vcenter','fg_color': '#FF8080', 'font_color':'#800000'}) # blue
    
    
    #worksheet.merge_range('C5:G5', 'Pre Trade', merge_format1)
    worksheet.merge_range(str('C'+str(st_row)+':'+'G'+str(st_row)), 'Pre Trade', merge_format1)
    worksheet.merge_range(str('H'+str(st_row)+':'+'J'+str(st_row)), 'Flow', merge_format5)
    worksheet.write_string(str('K'+str(st_row)), 'Target', merge_format6)
    worksheet.merge_range(str('L'+str(st_row)+':'+'O'+str(st_row)), 'Trade', merge_format2)
    worksheet.merge_range(str('P'+str(st_row)+':'+'U'+str(st_row)), 'Post Trade', merge_format3)
    worksheet.merge_range(str('V'+str(st_row)+':'+'X'+str(st_row)),  'Sign-off', merge_format2)
    
    worksheet.write(7, 12, "No. Futures / Price", cell_format2_2)
    
    worksheet.set_column('A:A', 16)
    worksheet.set_column('B:B', 6)
    worksheet.set_column('C:C', 17)
    worksheet.set_column('D:D', 17)
    worksheet.set_column('E:E', 15)
    worksheet.set_column('F:F', 15)
    worksheet.set_column('G:G', 13)
        
    worksheet.set_column('H:H', 13)        
    worksheet.set_column('I:I', 12)        
    worksheet.set_column('J:J', 11)
    
    worksheet.set_column('K:K', 11)
    
    worksheet.set_column('L:L',11)
    worksheet.set_column('M:M',11)
    worksheet.set_column('N:N',11)
    worksheet.set_column('O:O',9)
    worksheet.set_column('P:P',15)
    worksheet.set_column('Q:Q',14)
    worksheet.set_column('R:R',12)
    worksheet.set_column('S:S',13)
    worksheet.set_column('T:T',14)
    worksheet.set_column('U:U',11)
    worksheet.set_column('V:V',10)
    worksheet.set_column('W:W',13)
    worksheet.set_column('X:X',11)

#    worksheet.set_column('F:F', 15, None, {'level': 1})
    # Get the xlsxwriter workbook and worksheet objects.
    #workbook  = writer.book
    #worksheet = writer.sheets['Sheet1']
    
    # Add some cell formats.
    format1 = workbook.add_format({'num_format': 'R#,##0'})
    format2 = workbook.add_format({'num_format': '0.000%'})
    format3 = workbook.add_format({'num_format': '0', 'font_color': 'black', 'bold':True})
    format4 = workbook.add_format({'num_format': '0', 'font_color': 'black', 'bold':False})
    format5 = workbook.add_format({'bold': 1,'bg_color': '#C6EFCE', 'font_color':'#006100'})
    format6 = workbook.add_format({'bold': 1,'bg_color': '#FFC7CE', 'font_color':'#9C0006'})
    format7 = workbook.add_format({'num_format': '0.000%','bold':True,'font_color': 'red' })
    format8 = workbook.add_format({'num_format': '0.000%','bold':True,'font_color': 'green' })
    
    # Note: It isn't possible to format any cells that already have a format such
    # as the index or headers or any cells that contain dates or datetimes.
    
    # Set the column width and format.
    #worksheet.set_column('B:B', 18, format1)
    
    # Set the format but not the column width.
    #worksheet.set_column('C:C', None, format2)
    len1=n_comb_eff_1.shape[0]
    for i in range(st_it,len1+st_it,2):
   #    print(i)
       worksheet.set_row(i, 18, format1)
    
    for j in range(st_it,len1+st_it,2):
 #      print(i+1)
       worksheet.set_row(j+1, 18, format2)
    
      
    worksheet.conditional_format(str('M'+str(st_row+2)+':'+'M100'), {'type': 'cell',
                                                                     'criteria': 'between',
                                                                     'minimum': 1,
                                                                     'maximum': 9999,
                                                                     'format': format3})
    worksheet.conditional_format(str('M'+str(st_row+2)+':'+'M100'), {'type': 'cell',
                                                                     'criteria': '>',
                                                                     'value': 10000,
                                                                     'format': format1})
    worksheet.conditional_format(str('M'+str(st_row+2)+':'+'M100'), {'type': 'cell',
                                                                     'criteria': '=',
                                                                     'value': 0,
                                                                     'format': format4})
    worksheet.conditional_format(str('M'+str(st_row+2)+':'+'M100'), {'type': 'cell',
                                                                     'criteria': '<',
                                                                     'value': 0,
                                                                     'format': format3})
  
    worksheet.conditional_format(str('O'+str(st_row+2)+':'+'O100'), {'type': 'cell',
                                                                     'criteria': '=',
                                                                     'value': '"Buy"',
                                                                     'format': format5})
    worksheet.conditional_format(str('O'+str(st_row+2)+':'+'O100'), {'type': 'cell',
                                                                     'criteria': '=',
                                                                     'value': '"Sell"',
                                                                     'format': format6})
    
           
    for i in range(1, len(n_comb_eff_1),2):
   #     print(i+6)
        if n_comb_eff_1['Effectivecash_R'].iloc[i] <= n_comb_eff_1['Min_EffCash'].iloc[i]:
            worksheet.write(i+st_it, 9, n_comb_dta['Effectivecash_R'].iloc[i], format7)
        elif n_comb_eff_1['Effectivecash_R'].iloc[i] >= n_comb_eff_1['Max_EffCash'].iloc[i]:
            worksheet.write(i+st_it, 9, n_comb_dta['Effectivecash_R'].iloc[i], format7)
    
        if n_comb_eff_1['Totalcash_R'].iloc[i] <= n_comb_eff_1['Min_TotalCash'].iloc[i]:
            worksheet.write(i+st_it, 8, n_comb_dta['Totalcash_R'].iloc[i], format7)
            worksheet.write(i+st_it, 17, n_comb_dta['Totalcash_TR'].iloc[i], format7)
            if (n_comb_eff_1['Totalcash_R'].iloc[i]!=0):
                worksheet.write(i+st_it, 20, n_comb_dta['Check cash'].iloc[i], format6)
            else:
                worksheet.write(i+st_it, 20, '',cell_format1_1)
            
        elif n_comb_eff_1['Totalcash_R'].iloc[i] >= n_comb_eff_1['Max_TotalCash'].iloc[i]:
            worksheet.write(i+st_it, 8, n_comb_dta['Totalcash_R'].iloc[i], format7)
            worksheet.write(i+st_it, 17, n_comb_dta['Totalcash_TR'].iloc[i], format7)
           # worksheet.write(i+st_it, 20, n_comb_dta['Check cash'].iloc[i], format6)
            if (n_comb_eff_1['Totalcash_R'].iloc[i]!=0):
                worksheet.write(i+st_it, 20, n_comb_dta['Check cash'].iloc[i], format6)
            else:
                worksheet.write(i+st_it, 20, '',cell_format1_1)
            
        if n_comb_eff_1['Effectivecash_TR'].iloc[i] <= n_comb_eff_1['Min_EffCash'].iloc[i]:
            worksheet.write(i+st_it, 19, n_comb_dta['Effectivecash_TR'].iloc[i], format7)
        elif n_comb_eff_1['Effectivecash_TR'].iloc[i] >= n_comb_eff_1['Max_EffCash'].iloc[i]:
            worksheet.write(i+st_it, 19, n_comb_dta['Effectivecash_TR'].iloc[i], format7)
        else:
            worksheet.write(i+st_it, 19, n_comb_dta['Effectivecash_TR'].iloc[i], format8)
            
        if (n_comb_eff_1['Inflow'].iloc[i] < 0)&(~np.isnan(n_comb_eff_1['Inflow'].iloc[i])) :
            worksheet.write(i+st_it, 7, n_comb_dta['Inflow'].iloc[i], format7)
        elif (n_comb_eff_1['Inflow'].iloc[i] > 0)&(~np.isnan(n_comb_eff_1['Inflow'].iloc[i])):
            worksheet.write(i+st_it, 7, n_comb_dta['Inflow'].iloc[i], format8)
       
        
        if n_comb_eff_1['Tgt_EffCash1'].iloc[i] != n_comb_eff_1['Tgt_EffCash'].iloc[i]:
            worksheet.write(i+st_it, 10, n_comb_dta['Tgt_EffCash1'].iloc[i], merge_format4)
       
           
    # Close the Pandas Excel writer and output the Excel file.
    worksheet.autofilter(str('A'+str(st_it)+':'+'X100'))
    worksheet.set_row(st_row, 29.25)
#    worksheet.write(st_row, 1, 'Portfolio code', cell_format3)
#    worksheet.protect()
    
  #  worksheet.set_column('R:R', None, unlocked)
    
    for j in range(1, len(n_comb_eff_1),2):
 #       print(n_comb_dta.index.values[j][0])
 #       if n_comb_eff_1['Check cash'].ix[j] == '':
         worksheet.write(j+st_it, 21, '', cell_format3)
         worksheet.write(j+st_it, 22, '', cell_format3)
         worksheet.write(j+st_it, 23, '', cell_format3)
      #   worksheet.write(str('A'+str(j+st_it)), n_comb_dta.index.values[j][0], format3)
     #   worksheet.write_formula(str('A'+str(j+st_it)), str('='+str('A'+str(j+st_it))))  
   #     else:
   #         worksheet.write(j+st_it, 18, 0, cell_format4)
            #worksheet.write(str('R'+str(j+7)), '',  cell_format3)
  #      worksheet.write(str('R'+str(j)), '',  cell_format4)    

       # print(str('R'+str(j+6)))
    
 #   cell.protection = Protection(locked=False)
    writer.save()
    workbook.close()
    
       
    mywb = px.load_workbook(output_file)

    mysheet = mywb.active
    border1=Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='hair'), 
                     bottom=Side(style='hair'))
   # top = Border(top=border.top)
    
    
    for j in range(1, len(n_comb_eff_1),2):
     #  print(str('A'+str(j+st_it))+":"+str('A'+str(j+st_it+1)))
       mysheet.unmerge_cells(str('A'+str(j+st_it))+":"+str('A'+str(j+st_it+1)))
       mysheet.cell(row = j+st_it+1, column = 1).value = str("="+str('A'+str(j+st_it)))
       mysheet.cell(row = j+st_it+1, column = 1).font=Font(color=colors.WHITE)
       mysheet.cell(row = j+st_it+1, column = 1).border = border1
       
 #   mysheet.protection.sheet = True   
 #   area =mysheet['A7':'Q50']
 #   area.protection = Protection(locked=False)
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Trade at spot,Trade at close, Do not trade"', allow_blank=True)
    dv.error ='Your entry is not in the list'
    dv.errorTitle = 'Invalid Entry'
    dv.prompt = 'Please select from the list'
    dv.promptTitle = 'List Selection'
    mysheet.add_data_validation(dv)
    dv.add(str('W'+str(st_it+1))+":"+str('W'+str(st_it+100)))
    
    from openpyxl.worksheet.datavalidation import DataValidation
    dw = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    dw.error ='Your entry is not in the list'
    dw.errorTitle = 'Invalid Entry'
    dw.prompt = 'Please select from the list'
    dw.promptTitle = 'List Selection'
    mysheet.add_data_validation(dw)
    dw.add(str('V'+str(st_it+1))+":"+str('V'+str(st_it+100)))
   
    
    
    mysheet.protection.sheet = True
    mysheet.protection.password = 'Flower'
    mysheet.protection.autoFilter=False
 #   for h in range(1, len(n_comb_eff_1),1):
 #       cell = mysheet[str('A'+str(h))]
 #       cell.protection = Protection(autofilter=True, locked=False)
    from openpyxl.comments import Comment
    comment = Comment('Sign-off','PM')
    comment.height = 20
    mysheet["B6"].comment = comment
    mysheet.sheet_view.zoomScale = 84
    mywb.save(output_file) 
    
    os.startfile(output_folder)
"""
'******************************************************************************************************************************************************************************    
'                                                     Define user input to check if flows are updated
'******************************************************************************************************************************************************************************    
"""
    
def input_fx(termi_nate_cnt=5):
    cnt=0
    loop=True
    while loop:
        d1a = input ("1. Are flows & cash limits up to date: 1) Yes. 2) No.[Y/N]?: ")
        
        if d1a=="Y":
            print ("Futures report generation in progress",end='', flush=True)
            break
        elif d1a=="N":
            print ("Please update flows file",end='', flush=True)
            break
        else:
            cnt=cnt+1
            #print(cnt)
            print("Invalid input, please select the correct option")
            if cnt==termi_nate_cnt:
                print("You have run out of options, default option selected")
                d1a='N'
                #loop=False
                #print(loop)
                break
    
    if loop:
        x = [d1a]

    return x    
"""
'******************************************************************************************************************************************************************************    
'                                                     Define futures load function for futures tloader to Decalog
'******************************************************************************************************************************************************************************    
"""

# Tloader for futures

def tloader_fmt_futures(termi_nate_cnt=5):

    import sys
    import pandas as pd
    import numpy as np
    import datetime as dt
    from datetime import datetime, timedelta
    import glob
    import os
    #from pydatastream import Datastream
    #from business_calendar import Calendar, MO, TU, WE, TH, FR
    import pyodbc
    import xlrd
    from tkinter import filedialog
    from tkinter import Tk

        
    startDate = datetime.today().date()
    
    folder_yr = datetime.strftime(startDate, "%Y")
    folder_mth = datetime.strftime(startDate, "%m")
    folder_day = datetime.strftime(startDate, "%d")
    
    
    dirtoimport_file='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES'
    
    #output_folder='\\'.join([dirtooutput_file ,folder_yr, folder_mth])
    input_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\Futures Trades\\')
    
    root = Tk()
    root.filename =  filedialog.askopenfilename(initialdir = input_folder,title = "choose your file",filetypes = (("*.xlsx","*.xlsx"),("all files","*.*")))
  #  print (root.filename)
    root.withdraw()
    
    if root.filename=='':
        msg="No file selected!"
    else:
    
    #    dic_users={'blala':['BLL','blala'], 'test':['TST','test'], 'sbisho':['SB','sbisho'], 'tmfelang2':['TM','tmfelang'], 'abalfour':['AB','abalfour'], 'sparker2':['SP','sparker'], 'fsibiya':['FS','fsibiya']}
        user_dict=pd.read_csv('C:\\IndexTrader\\required_inputs\\user_dictionary.csv')
        dic_users=user_dict.set_index(['username']).T.to_dict('list')
  
       # dirtooutput_file= 'U:\\Production\\In\\'
       # dirtooutput_file= 'c:\\data\\'
        dirtooutput_file = '\\\\za.investment.int\\DFS\\SSDecalogUmbono\\Production\\In\\'   
        
        #newest = max(glob.iglob(input_folder+'IndexFutRep_*.xlsx'), key=os.path.getmtime)
        
        #check = xlrd.open_workbook(newest)
        check = xlrd.open_workbook(root.filename)
        sh='Sheet1'
        sht_nam=check.sheet_names()    
        
        ls = []
        cols = range(0,23)
        for i in cols:
            #    print(i)
            ls.append(i)
        
        #fund_xls = pd.read_excel(newest, sheet_name = sh, skiprows =7, usecols = ls)
        fund_xls = pd.read_excel(root.filename, sheet_name = sh, skiprows =7, usecols = ls)
        
        if 'Sign-Off' in sht_nam:
            print(sht_nam)
            checksheet = check.sheet_by_name('Sign-Off')
            if checksheet.nrows > 5:
                value_app = checksheet.cell_value(5,0) 
            else: 
                value_app =''
            
        else:
         #   msg='Please sign off trade first!'
            value_app=''
           
        
        checksheet2 = check.sheet_by_name('Sheet1')
        
        run_load=1
        for j in range(9,(len(fund_xls)+9),2):
            value = checksheet2.cell(j, 20)
            if str(value) == 'number:1.0':
                value_cell = checksheet2.cell(j, 21)
                if str(value_cell)[0:11]=="text:'Trade":
                  run_load_x=1    
                else:
                  run_load=0
                #  print("Please enter a trade ctloader_fmt_futuresomment before loading")
             #        break
      
        if run_load==1:
            
            if (str(value_app) in ['Approved',"text:'Approved'"]):
            
                fund_xls_ex= fund_xls[fund_xls.TradeComment.isin(['Trade at spot','Trade at close'])]
                fund_xls_ex= fund_xls_ex[fund_xls_ex.TradeSignal.isin([1])]
                fund_xls_ex=fund_xls_ex.copy()
                fund_xls_ex['TradeShort']= np.where(fund_xls_ex['Trade'].values=='Sell', 'SOC', 
                                                             np.where(fund_xls_ex['Trade'].values=='Buy', 'BOC', ''))
                fund_xls_ex['Nom']=abs(fund_xls_ex['No. Futures / Price'].values)
                fund_xls_x=fund_xls_ex[['FutureCode', 'Portfolio Code', 'TradeShort', 'Nom']]
                fund_xls_x=fund_xls_x.copy()
                fund_xls_x['Instruction']='Rebalance Portfolio'
                fund_xls_x['MP']='MP'
                fund_xls_x['Blank']=''
                fund_xls_x['TradeIns']=  fund_xls_ex['TradeComment']
                
                
                with open(str(dirtooutput_file+"FuturesTrade"+folder_yr+folder_mth+folder_day+'.txt'), "w") as fin:
              #  with open(str('c:\\data\\'+"FuturesTrade"+folder_yr+folder_mth+folder_day+'.txt'), "w") as fin:
                    #fin.write('\n'.join((fund_xls_ex.values.tolist())[0]))
                    for i in range(0,len(fund_xls_ex)):
                       # print(i)
                        st=fund_xls_x.values.tolist()[i]
                        sf=st[0]+','+st[1]+',',st[2]+','+str(int(st[3]))+','+st[4]+','+st[5]+','+st[6]+','+st[7]+'\n'
                        sh=''.join(sf)
                        #sf=st[0]+','+st[1]+',',st[2]+','+st[4]+','+st[5]+','+st[6]+','+st[7]
                        fin.write(sh)
                    # Get sheet names
                os.startfile(dirtooutput_file)  
                print("Trades loaded into Decalog!!")
                msg="Trades loaded into Decalog!!"
                    
            else:
                print("Trade has not been approved, \nplease sign-off before loading, \n Trades are not loaded!!")
                msg="Trade has not been approved, \nplease sign-off before loading, \nTrades are not loaded!!"
                
        else:
            print("Please enter a trade comment before loading, Trades are not loaded!!")
            msg="Please enter a trade comment before loading,\nTrades are not loaded!!"
    return msg

"""
        
'******************************************************************************************************************************************************************************    
'                                                     Define input function for Equity tloader to Decalog
'******************************************************************************************************************************************************************************    
"""
        

def tloader_fmt_equity(selct_on=1,SSF_Funds=['CORPEQ','MFEQTY'], SSF_parent=['OMU']):

    import sys
    import pandas as pd
    import numpy as np
    import datetime as dt
    from datetime import datetime, timedelta
    import datetime as dt
    import glob
    import os
    #from pydatastream import Datastream
    #from business_calendar import Calendar, MO, TU, WE, TH, FR
    import pyodbc
    import xlrd
    from tkinter import filedialog
    from tkinter import Tk
    from tkinter import messagebox
    import tkinter as tk
    from tkinter.simpledialog import askstring
    from tkinter.messagebox import showinfo
    
    
    from write_excel import special_fund_load as sp_acc_fnd
    from write_excel import cash_fx_pre_trd_comp as cash_fx_pre_trd_comp
    
    
    startDate = datetime.today()
    
    folder_yr = datetime.strftime(startDate, "%Y")
    folder_mth = datetime.strftime(startDate, "%m")
    folder_day = datetime.strftime(startDate, "%d")
    
    user_dict=pd.read_csv('C:\\IndexTrader\\required_inputs\\user_dictionary.csv')
    dic_users=user_dict.set_index(['username']).T.to_dict('list')
        
    
        
    dirtoimport_file='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES'
    user_dta='\\\\za.investment.int\\dfs\\dbshared\\OMGxT\\Aegis\\UserData.xls'
    
    #output_folder='\\'.join([dirtooutput_file ,folder_yr, folder_mth])
    input_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\BatchTrades\\'+dic_users[os.environ.get("USERNAME").lower()][1]+'\\')
   # input_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day]))
    
    
    
 #   dic_users={'blala':['BLL','blala'], 'test':['TST','test'], 'sbisho':['SB','sbisho'], 'tmfelang2':['TM','tmfelang'], 'abalfour':['AB','abalfour'], 'sparker2':['SP','sparker'], 'fsibiya':['FS','fsibiya']}
    user_dict=pd.read_csv('C:\\IndexTrader\\required_inputs\\user_dictionary.csv')
    dic_users=user_dict.set_index(['username']).T.to_dict('list')
  
    #dirtooutput_file= 'U:\\Production\\In\\'
    #dirtooutput_file='c:\\data\\'
    dirtooutput_file = '\\\\za.investment.int\\DFS\\SSDecalogUmbono\\Production\\In\\'   
    
    #newest = max(glob.iglob(input_folder+'Trade*.xlsx'), key=os.path.getmtime)
    
    sh='TradeList'
    unlst_sec = ['JCDLIQ','ZAEHSU','ZAUTRU']
    chck_shts = ['Post_Trade_Report', 'Active_Bets', 'Raw_Data', 'Sign-Off']
    
    ls = []
    cols = range(0,20)
    for i in cols:
 #       print(i)
        ls.append(i)
    
    print(selct_on)
    if selct_on>0:
        messagebox.showinfo('Trade Imports', 'Please select the the APPROVED Post Trade Recon Report you uploading trades for!!' )   
        root = Tk()
        root.filenamePT =  filedialog.askopenfilename(initialdir = str(input_folder),title = "Choose your Post Trade Recon file",filetypes = (("all files","*.csv"),("all files","*.*")))
      #  print (root.filename)
        root.withdraw()
        if root.filenamePT=='':
            selct_on=4
    else:
        selct_on=0
    
    if selct_on==0:
       #selct_on=4
       msg1=''
       value_app=''
       value_date=''
    else:
            
        if selct_on==4:
       #     selct_on=4
            msg1='Please select approved \n Post Trade Recon file'
            value_app=''
            value_date=''
        else:
            try:
                check = xlrd.open_workbook(root.filenamePT)
                sh='Sheet1'
                sht_nam=check.sheet_names()
                cr_book=all(f in chck_shts  for f in sht_nam)
                value_app=''
                value_date=''
                
                if cr_book:
                    
                    sh_f_c=check.sheet_by_name('Post_Trade_Report')
                    
                    if 'Sign-Off' in sht_nam:
                         print(sht_nam)
                         checksheet = check.sheet_by_name('Sign-Off')
                         if checksheet.nrows > 5:
                             value_app = checksheet.cell_value(5,0)
                             value_date = dt.datetime(*xlrd.xldate_as_tuple(checksheet.cell_value(5,2),check.datemode)).date()
                             fnd_lst=[]
                             for j in range(3,60,2):
                           #      print(j)
                                 try:
                                     get_fnds=sh_f_c.cell_value(2,j)
                                     fnd_lst.append(get_fnds)
                                 except: 
                                     pass
                                                 
                         else: 
                             value_app =''
                             value_date=''
                             selct_on=4
                        
                    else:
                        value_app=''
                        value_date=''
                        selct_on=4
            except:
                    messagebox.showwarning("Warning","You have selected the incorrect file!!\nPlease try again!")
                    msg1='Please load \nthe correct file'
                    value_app=''
                    value_date=''
                    selct_on=4
                
        if (value_app!='Approved'):
            msg1='Please load an\napproved Post\nTrade Recon file'
            selct_on=4
           
        else:
            if value_date!=startDate.date():
                msg1='Please load an\n Post Trade Recon file \n with correct date'
                selct_on=4
            else:
            
                if selct_on == 1:
                
                    root = Tk()
                    root.filename =  filedialog.askopenfilename(initialdir = str(input_folder+'Equity Trades'),title = "Choose your Equity file",filetypes = (("*.csv","*.csv"),("all files","*.*")))
              #  print (root.filename)
                    root.withdraw()
                    
                elif selct_on == 3:
                    
                    root = Tk()
                    root.filename =  filedialog.askopenfilename(initialdir = str(input_folder+'Equity Trades'),title = "Choose your Equity file",filetypes = (("*.csv","*.csv"),("all files","*.*")))
                    #root.withdraw()
                    
                   # root2 = Tk()
                    root.filenameF =  filedialog.askopenfilename(initialdir = str('\\'.join([input_folder])+'FuturesFile'),title = "Choose your Futures file",filetypes = (("*.csv","*.csv"),("all files","*.*")))
              #  print (root.filename)
                    root.withdraw()
                    
                elif selct_on == 2:
                    root = Tk()
                    root.filenameF =  filedialog.askopenfilename(initialdir = str('\\'.join([input_folder])+'FuturesFile'),title = "Choose your Futures file",filetypes = (("*.csv","*.csv"),("all files","*.*")))
                   
                else:
                    msg1='No trades to \n be loaded!'
                    
                run_job = 1
                if selct_on == 3:
                    if ((root.filename == '') or (root.filenameF=='')):
                        msg1="No trades loaded,\n no files selected"
                        run_job = 0
                        selct_on=4
                elif selct_on == 1:
                    if root.filename== '':         
                        msg1="No trades loaded,\n no equity file selected"
                        run_job = 0
                        selct_on=4
                elif selct_on == 2:
                    if root.filenameF== '':         
                        msg1="No trades loaded,\n no futures file selected"
                        run_job = 0
                        selct_on=4
                else:
                    print("Run job")
                    
                if run_job==1:
                    
                    if selct_on in [1,3]:
                        chck_trdlist=['Date', 'Initial Portfolio Name', 'Asset ID', 'Asset Name','Initial Holdings', 'Trade', 
                                                'Final Holdings', 'Price', 'Traded Value','Final Value', 'Trade Type']
                        fund_xls = pd.read_csv(root.filename, skiprows =1, header = 0)
                        ck_e_cf=all(t_l in chck_trdlist for t_l in fund_xls.columns)
                        
                        if ck_e_cf==True:
                        
                            print(len(fund_xls))
                            fund_xls = fund_xls.drop_duplicates(['Initial Portfolio Name','Asset ID','Initial Holdings','Trade','Final Holdings','Price','Traded Value','Final Value','Trade Type'],keep='first')
                            print(len(fund_xls))
                            fnd_eq=fund_xls['Initial Portfolio Name'].unique().tolist()
                         #   print(fnd_eq)
                            fnd_eq=[j for j in fnd_eq if str(j) != 'nan']
                            y_fund=(all(x in fnd_lst for x in fnd_eq)) # checks if funds in equity trade list is in fund post opt comb recon report
                            print("pass through here")
                        else:
                            messagebox.showerror("Incorrect Equity Trade File", "You have selected the incorrect file")
                            msg1="No load,\n Incorrect equity\nfile selected"
                            selct_on=4
                            run_job=0
                            y_fund=False
                            
                         
                    elif selct_on==2:
                        fut=pd.read_csv(root.filenameF, header = None) 
                        fut_port=fut[1].tolist()
                        y_fund=(all(x in fnd_lst for x in fut_port)) 
                        
                    if y_fund:
                        
                        if selct_on in [1,3]:
                            #fund_xls['AlpCode']= (fund_xls['Asset ID'])[1:]
                            fund_xls = fund_xls[fund_xls['Asset ID'] != 'ZAR']
                            fund_xls_chc= (fund_xls.drop_duplicates(['Initial Portfolio Name','Asset ID'],keep='first')).shape[0]-fund_xls.shape[0]
                            if fund_xls_chc==0:
                                #print(fund_xls_chc)
                                fund_xls.loc[:,'AlpCode'] = fund_xls['Asset ID'].apply(lambda x : x[2:] if x.startswith("ZA") else x)   
                                fund_xls.loc[:,'TradeShort']= np.where(fund_xls['Trade Type'].values=='SELL', 'S', 
                                                                             np.where(fund_xls['Trade Type'].values=='BUY', 'B', ''))
                                fund_xls.loc[:,'Short_sell']=np.where(fund_xls['Final Holdings']<0, 1,0)
                                kk=fund_xls.groupby(['Initial Portfolio Name']).agg({'Traded Value':'sum'}).reset_index()
                                
        
                                
                                if any(g in fund_xls['Asset ID'].unique() for g in unlst_sec):
                                    messagebox.showwarning("Warning","Trading unlisted shares!!\nPlease check trades!")
                                    msg1='Trades not loaded,\n unlisted shares!'
                                    selct_on = 4
                                    run_job =0
                                else:    
                                    
                                    if (fund_xls[fund_xls.Short_sell.isin([1])]).shape[0]==0:
                                        fund_xls=fund_xls.drop(['Short_sell'], axis=1)
                                        fund_xls.loc[:,'BloomCode']= fund_xls['AlpCode'].apply(lambda x: "{}{}".format(x, ' SJ'))
                                    
                                        user_dat=pd.read_excel(user_dta, usecols = ls)
                                        user_dat=user_dat[["!ID","'BB_TICKER","MIN_AVG_VOLUME"]]
                                        user_dat.columns=['ID','BB_Ticker','VOL']
                                        user_dat=user_dat.drop_duplicates(['BB_Ticker'], keep='first')
                                        fund_xls=pd.merge(fund_xls, user_dat, left_on=["BloomCode"], right_on=['BB_Ticker'], how="left")
                                        fund_xls.loc[:,"TradeDays"]=np.abs(fund_xls["Trade"].values)/(0.2*fund_xls["VOL"].values)
                                        fund_xls.loc[:,"TradeDays"]=(fund_xls["TradeDays"]).fillna(0)
                                        fund_xls.loc[:,'TradeAction']= np.where(fund_xls['TradeDays'].values<0.3, 'Trade at close', 
                                                                                     np.where(fund_xls['TradeDays'].values < 0.8,'Target close', 
                                                                                              'Trade in line with market'))
                                        
                                        
                                        csh_chck=cash_fx_pre_trd_comp(fnds_to_use=fnd_eq,response='yes',orders=True,testing=False)
                                        all_dt=csh_chck[0]
                                        all_dt=pd.merge(all_dt,kk, left_on=['Port_code'] ,right_on=['Initial Portfolio Name'], how='right')
                                        all_dt.loc[:,'CashBal']=all_dt['Settled_cash'].values-all_dt['Traded Value'].values
                                        ldt_cf=[]
                                        for f_d in fnd_eq:
                                            val_ =all_dt[all_dt.Port_code==f_d].CashBal.values[0]
                                            if val_ <0:
                                                ldt_cf.append(str(f_d+':'+' Neg Cash Bal: '+str('R {:10,.2f}'.format(val_))))
                                            else:
                                                pass
                                        if len(ldt_cf)>0:
                                           inp1=messagebox.askyesno('Warning (Overdraft):', str('Please note following funds will go into overdraft: \n'+'\n'.join(map(str, ldt_cf))+'\nWould you like to continue?' ) ) 
                                        else:
                                            inp1=True
                                        
                                        if inp1==False:
                                           messagebox.showinfo('Warning (Overdraft):', str('Please note you have terminated Trade laod process!!') )
                                           msg1='Trades not loaded,\n due to overdraft!'
                                           selct_on = 4
                                           run_job =0
                                            
                                        else:
                                           
                                            inp=messagebox.askyesno("Change equity trade instruction", "Would you like to enter an alternative trade instruction?\n E.g. Trade at spot 11am; Trade at spot")      
                                            
                                            if inp:                                       
                                                foot = tk.Toplevel()
                                                foot.geometry(str("270x"+"110"+"+600+400"))
                                                foot.title('Equities Trade Instruction')
                                                
                                                tk.Label(foot, 
                                                         text="Please enter equity trade instruction:", font="Helvetica 10").grid(row=0, pady=5,padx=25)
                                                
                                                d1 = tk.Entry(foot, width=30)
                                                
                                                d1.grid(row=1, column=0, pady=5)
                                                
                                                tk.Button(foot, 
                                                          text='OK', 
                                                          command=foot.quit).grid(row=2,  column=0, pady=5
                                                                                  
                                                                                    )
                                                foot.mainloop()
                                                foot.withdraw()
                                                fund_xls.loc[:,'TradeAction']=np.where(fund_xls['TradeAction']=='Trade at close', str(d1.get()),fund_xls['TradeAction'].values)
                                            else:
                                                pass
                                            if (any(pd.Series(fund_xls['Initial Portfolio Name'].unique()).isin(SSF_Funds))&(any(pd.Series(fund_xls['AlpCode'].unique()).isin(SSF_parent)))): 
                                                
                                                 inp_ssf=messagebox.askyesno("Tradind Single Stock Future", str(str("Would you like to import SSF or underlying share for ")+SSF_parent[0]+str('?\n[Yes] SSF; [No] Share')))      
                                                 try:
                                                     fut_suf=csh_chck[1]
                                                 except:
                                                     fut_suf = askstring('SSF Details', 'Please enter the SSF Code \ne.g OMUFS19?')
                                                     showinfo("Load SSF", "You're about to load {}".format(fut_suf))
                                                     fut_suf = fut_suf[:-3]
                                                 
                                                 if inp_ssf:
                                                     def f(x):
                                                         return np.int(x)
                                                     f2 = np.vectorize(f)
                                                     fund_xls.loc[:,'Trade']=np.where(fund_xls['Initial Portfolio Name'].isin(SSF_Funds), 
                                                                                        np.where(fund_xls['AlpCode'].isin(SSF_parent), f2(fund_xls['Trade'].values/100),fund_xls['Trade'].values),
                                                                                        fund_xls['Trade'].values)
                                                     fund_xls.loc[:,'TradeShort']=np.where(fund_xls['Initial Portfolio Name'].isin(SSF_Funds), 
                                                                                        np.where(fund_xls['AlpCode'].isin(SSF_parent),(fund_xls.TradeShort+'OC'), fund_xls['TradeShort'].values),
                                                                                        fund_xls['TradeShort'].values)
                                                     fund_xls.loc[:,'AlpCode']=np.where(fund_xls['Initial Portfolio Name'].isin(SSF_Funds), 
                                                                                        np.where(fund_xls['AlpCode'].isin(SSF_parent),(fund_xls.AlpCode+'F'+fut_suf), fund_xls['AlpCode'].values),
                                                                                        fund_xls['AlpCode'].values)
                                                     
                                                 else:
                                                     pass
                                            inp_ta=messagebox.askyesno('Trading/Special Accounts:', str('Would you like to load to any trading or special accounts?' ) )  
                                            
                                            if inp_ta:
                                                
                                                acc_o_load=sp_acc_fnd(fnd_eq,ans=inp_ta, fnd_dir='C:\\IndexTrader\\required_inputs\\', fnd_dict='fund_dictionary.csv')[0]
                                                fund_xls.loc[:,'Initial Portfolio Name']=fund_xls['Initial Portfolio Name'].map(lambda x:acc_o_load[x])
                                            else:
                                                pass
                                                
                                    else:    
                                        messagebox.showwarning("Warning","Short-selling shares!!\nPlease check trades!")
                                        msg1='Trades not loaded,\nshort-selling!'
                                        selct_on = 4
                                        run_job =0
                            else:
                                messagebox.showwarning("Warning","Duplicate trades detected, please check!")
                                msg1='There are duplicate trades! \nPlease check your file!!'
                                selct_on = 4
                                run_job =0
                          #  return msg1
                            #break
          #      if run_job ==1:
                    else:
                        messagebox.showwarning("Warning","Please select the correct funds to load!")
                        msg1='Please select the correct funds \n you want to load!'
                        selct_on=4
                        
                        
                #     fund_xls_ex= fund_xls.loc[fund_xls['Trade Comment'] == 1]
                #    fund_xls_ex['Nom']=abs(fund_xls_ex['No. Futures'].values)
                #    fund_xls_ex=fund_xls_ex[['FutureCode', 'Portfolio Code', 'TradeShort', 'Nom']]
                #    fund_xls_ex['Instruction']='Rebalance Portfolio'
                #    fund_xls_ex['MP']='MP'
                #    fund_xls_ex['Blank']=''
                #    fund_xls_ex['TradeIns']='Trade at spot'
                    if (((y_fund)&(selct_on==3))|(selct_on==2)): 
                                                
                        fut_trdlist=['OC', 'MP']
                        try:
                            fut=pd.read_csv(root.filenameF, header = None) 
                            fk_e_cf=all(t_f in fut_trdlist[1] for t_f in fut.iloc[:,5].values[0])&(all(d_f in fut_trdlist[0] for d_f in fut.iloc[:,2].values[0][1:]))
                        except:
                            fk_e_cf=False
                        
                        if fk_e_cf==True:
                    
                            fut[6]=''
                            fut[3]= fut[3].astype(int)
                            #messagebox.showinfo("Futures","You are trading futures")
                            
                            #master = tk.Tk()
                            coot = tk.Toplevel()
                            coot.geometry(str("270x"+"110"+"+600+400"))
                            coot.title('Futures Instruction')
                                  
                            tk.Label(coot, 
                                     text="Please enter futures trade instruction:", font="Helvetica 10").grid(row=0, pady=5,padx=25)
                            
                            e1 = tk.Entry(coot, width=30)
                            
                            e1.grid(row=1, column=0, pady=5)
                            
                            tk.Button(coot, 
                                      text='OK', 
                                      command=coot.quit).grid(row=2,  column=0, pady=5
                                                              
                                                                )
                            coot.mainloop()
                            coot.withdraw()
                            fut[7]=str(e1.get())
                        
                        else:
                            messagebox.showerror("Incorrect Futures File", "You have selected the incorrect futures file")
                            msg1="No load,\n Incorrect futures\nfile selected"
                            selct_on=4
                            run_job=0
                    else:
                        #msg1='Please sign-off trades \n you want to load!'
                         pass  
                    
                    
                    if selct_on in [1,2,3]:
                        p=messagebox.askquestion('Decalog Order Load', "All checks are complete!! Are you sure you want to load orders to Decalog?")
                        if p=='no':
                            selct_on=4
                            messagebox.showinfo('Decalog Order Load', "Order load cancelled.\nNo orders loaded to Decalog!")
                        else:
                            pass
                    else:
                        p=messagebox.showinfo('Trade Load to Decalog', "Some checks have failed.\nNo trades loaded to Decalog!")
                        selct_on=4
                    if selct_on in [1,2,3]:
                        with open(str(dirtooutput_file+"EquityTrade"+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.txt'), "w") as fin:
                     #   with open(str('c:\\data\\'+"EquityTrade"+folder_yr+folder_mth+folder_day+'.txt'), "w") as fin:
                            #fin.write('\n'.join((fund_xls_ex.values.tolist())[0]))
                            if ((selct_on in [1,3])&(y_fund)):
                                for i in range(0,len(fund_xls)):
                            #    for i in range(0,10):
                                #    print(i)
                                    st=fund_xls.values.tolist()[i]
                                    sf=st[11]+','+st[1]+','+st[12]+','+str(int(abs(st[5])))+',''Rebalance Portfolio'+','+'MP'+',,'+st[18]+'\n'
                                    sh=''.join(sf)
                                    #sf=st[0]+','+st[1]+',',st[2]+','+st[4]+','+st[5]+','+st[6]+','+st[7]
                                    fin.write(sh)
                                if selct_on==3:     
                                    for z in range(0,len(fut)):
                                        ft=','.join(str(e) for e in fut.values.tolist()[z])
                                        fh=ft+'\n'
                                        fin.write(fh)
                                        msg1="Futures and \n equities loaded"
                                else:
                                    msg1="Equities loaded only"
                                    
                            elif selct_on==2:
                                    for z in range(0,len(fut)):
                                        ft=','.join(str(e) for e in fut.values.tolist()[z])
                                        fh=ft+'\n'
                                        msg1="Futures loaded only"
                                        fin.write(fh)
                            else:
                                msg1="No load"
                        os.startfile(dirtooutput_file)  
                    else:
                        msg1="No load"
            
        if (msg1=="No load")&(selct_on in [1,2,3]):
            os.remove(str(dirtooutput_file+"EquityTrade"+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'))
        
        else:
            print("Happy!!")
        
#    msg1='No'
    return msg1
"""    
'******************************************************************************************************************************************************************************    
'                                                                   Select funds to trade
'******************************************************************************************************************************************************************************    
"""

def select_fund(struc=True,path_flow='C:\\IndexTrader\\required_inputs\\flows.csv'):
    import pandas as pd
    get_fund_list = pd.read_csv(path_flow,thousands=',')
    get_funds = (get_fund_list[(get_fund_list.Trade==1)])['Port_code'].tolist()
    if struc:
        get_funds = list(set(get_funds + ['CORPEQ'])) # Add corpeq to get the underlyng aset classifications    
    return get_funds



"""    
'******************************************************************************************************************************************************************************    
'                                                                   Check cash pre-trade compliance
'******************************************************************************************************************************************************************************    
"""

#def cash_fx_pre_trd_comp(lst_funds, csh_flow):

def cash_fx_pre_trd_comp(fnds_to_use=['Check'],response='yes',orders=True,testing=False):
        #import future
    
        import sys
        #sys.path.append('C:\Program Files (x86)\WinPython\python-3.6.5.amd64\lib\site-packages\IPython\extensions')
        #sys.path.append('C:\Program Files (x86)\WinPython\settings\.ipython')
        
        #for p in sys.path:
        #    print(p)
        
        import numpy as np
        
        import pandas as pd
        import time as tm
        import datetime as dt
        from datetime import datetime, timedelta
        import glob
        import os
        #from pydatastream import Datastream
        #from business_calendar import Calendar, MO, TU, WE, TH, FR
        import pyodbc
        #from write_excel import excel_fx as exl_rep
        #from write_excel import input_fx as inp
     #   from write_excel import select_fund as sf
        from write_excel import CashFlowFlag as cff
        from write_excel import trade_calc as t_c
        from write_excel import trade_calc_automatic as t_c_a
        from write_excel import bulk_cash_excel_report as bcer
        from write_excel import cash_flow_validity_fx as cfvf
        from write_excel import assetClassB as assetClass
        from write_excel import res_indB as res_ind
        from write_excel import fx_dtaB as fx_dta
        
        from bdateutil import isbday
        import bdateutil as bdut
        import holidays as hol        
        
        run_time = tm.time()
        if testing:
            response= 'yes'
            #automatic = True
            orders=False
            
        
        if response:
            start_time = datetime.now() 
            
            np.seterr(divide='ignore', invalid='ignore')
            
            """
            Set parameters for trading
            """
            startDate = datetime.today()
            pd.options.display.max_rows = 200
            dirtoimport_file='\\\\za.investment.int\\DFS\\SSDecalogUmbono\\IndexationPosFile\\'
            dirtoimport_cashfile = '\\\\za.investment.int\\DFS\\SSDecalogUmbono\\IndexationPosFile\\'
            #dirtoimport_cashfile = '\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES\\Decalog Valuation\\' 
            # Map fund and benchmark settings 
            
              # Pull in fund dictionary
            fnd_dict=pd.read_csv('C:\\IndexTrader\\required_inputs\\fund_dictionary.csv')
            dic_om_index=fnd_dict.set_index(['FundCode']).T.to_dict('list')
                
                
            user_dict=pd.read_csv('C:\\IndexTrader\\required_inputs\\user_dictionary.csv')
            dic_users=user_dict.set_index(['username']).T.to_dict('list')
                
            fnd_excp= ['DSALPC','OMCC01','OMCD01','OMCD02','OMCM01','OMCM02','PPSBTA','PPSBTB']
              
            override=['SSF DIV']             
            
            # Public Holidays
#            pub_holidays = (pd.read_excel("C:\\IndexTrader\\required_inputs\\public_holidays.xlsx"))['pub_holidays'].tolist()
            #cal = Calendar(holidays=pub_holidays)
            
                    
             # Determine list of funds to trade
            #lst_fund=sf(False)
            lst_fund=fnds_to_use
            
            
            # Import cash limits
            cash_lmt_x = pd.read_csv('C:\\IndexTrader\\required_inputs\\cash_limits.csv')
            cash_lmt_x=cash_lmt_x[cash_lmt_x.P_Code.isin(lst_fund)]
            cash_lmt_dict=cash_lmt_x.set_index(['P_Code'])[['Min_EffCash','Max_EffCash']].T.to_dict()
           
            
             
            
            # Import Flows
            #cash_flows_eff = pd.read_csv('H:\\Bernisha\\Work\\IndexTrader\\Data\\required_inputs\\flows.csv')
            cash_flows_eff = pd.read_csv('C:\\IndexTrader\\required_inputs\\flows.csv',thousands=',')
            cash_flows_eff=(cash_flows_eff[cash_flows_eff.Port_code.isin(lst_fund)]).drop('Trade',1)
            fut_sufx=(cash_flows_eff[['fut_sufx']]).values[0].tolist()[0]
            #print(fut_sufx)
            
            
             # Import futures
            fut_flow=pd.merge(cash_lmt_x[['P_Code','Future_Code']], cash_flows_eff[['Port_code','fut_sufx']], how='right', left_on=['P_Code'], right_on=['Port_code'] )
            fut_flow['Sec_code']= fut_flow[['Future_Code', 'fut_sufx']].apply(lambda x: ''.join(x), axis=1)
            fut_flow['Sec_code']=np.where(fut_flow.Future_Code=='NoFuture','NoFuture',fut_flow.Sec_code.values)
                
      
                    
            """
            Fund, Benchmark, Corporate Action data import
            """
            #newest = max(glob.iglob(dirtoimport_file+'fund_data/*.xls'), key=os.path.getmtime)
            newest = max(glob.iglob(dirtoimport_file+'*.xls'), key=os.path.getmtime)
            newest_cash=max(glob.iglob(dirtoimport_cashfile+'*.xls'), key=os.path.getmtime)
            #str(dirtoimport_file+newest)
            #newest
            
            fund_xls = pd.read_excel(newest,sheet_name=0,usecols='A:K',converters={'Portfolio':str, 'Price Date': pd.to_datetime, 
            'Inst Type':str, 
            'Inst Name':str,
            'ISIN':str,
            'Instrument':str,
            'Quote Close':float,
            'Qty':float,
            'Market Val':float,
            'Delta':float,	
            'Origin':str},
            )
    
            fund_xls=fund_xls.drop(['Delta'],axis=1)
            if orders:
                fund_xls.loc[:,'Inst Name']= np.where(fund_xls.Origin=='ORDER CASH', 
                                                  np.where(fund_xls['Inst Type'].str.split(" : ",n=1,expand=True)[0].values!="FUTURE", "DIF",fund_xls['Inst Name'].values),fund_xls['Inst Name'].values)
                fund_xls.loc[:,'Instrument']= np.where(fund_xls.Origin=='ORDER CASH', 
                                                  np.where(fund_xls['Inst Type'].str.split(" : ",n=1,expand=True)[0].values!="FUTURE", "ZAR",fund_xls['Instrument'].values),fund_xls['Instrument'].values)
                fund_xls.loc[:,'Inst Type']= np.where(fund_xls.Origin=='ORDER CASH', 
                                                  np.where(fund_xls['Inst Type'].str.split(" : ",n=1,expand=True)[0].values!="FUTURE","PAYABLE",fund_xls['Inst Type'].values),fund_xls['Inst Type'].values)
                fund_xls.loc[:,'Qty']= np.where(fund_xls.Origin=='ORDER CASH', 
                                                  np.where(fund_xls['Inst Type'].str.split(" : ",n=1,expand=True)[0].values=="FUTURE",0,fund_xls['Qty'].values),fund_xls['Qty'].values)
                fund_xls.loc[:,'Market Val']= np.where(fund_xls.Origin=='ORDER CASH', 
                                                  np.where(fund_xls['Inst Type'].str.split(" : ",n=1,expand=True)[0].values=="FUTURE",0,fund_xls['Market Val'].values),fund_xls['Market Val'].values)
         
            else:
                fund_xls=fund_xls[fund_xls.Origin=='POSITION']
            fund_xls=fund_xls.drop(['Origin'],axis=1)
            #fund_xls.dtypes
            
            fund_xls.columns = ['Port_code','Price_date','Sec_type','Sec_name','Sec_ISIN','Sec_code','Close_price','Quantity','Market_price']
            fund_xls['Close_price']=pd.to_numeric(fund_xls.Close_price.values, errors='coerce') 
            fund_xls['Quantity']=pd.to_numeric(fund_xls.Quantity.values, errors='coerce') 
            fund_xls['Market_price']=pd.to_numeric(fund_xls.Market_price.values, errors='coerce') 
            
            fund_obj = fund_xls.select_dtypes(['object'])
            fund_xls[fund_obj.columns] = fund_obj.apply(lambda x: x.str.strip())
            
            
            df=fund_xls.copy()
            df=df[(df.Port_code.isin(dic_om_index.keys()))]
                  
            df.loc[:,'Benchmark_code']=df.Port_code.map(lambda x:dic_om_index[x][1])
            df.loc[:,'TypeFund']=df.Port_code.map(lambda x:dic_om_index[x][2])
            
            df['Trade_date']=startDate
            df['AssetType1']=df.apply(lambda r: (assetClass(r.Sec_type,r.Sec_code, r.Sec_name,r.TypeFund,cash_flows_eff)).split(",")[0],axis=1)
            df['AssetType2']=df.apply(lambda r: (assetClass(r.Sec_type,r.Sec_code, r.Sec_name,r.TypeFund,cash_flows_eff)).split(",")[1],axis=1)
            df['AssetType3']=df.apply(lambda r: (assetClass(r.Sec_type,r.Sec_code, r.Sec_name,r.TypeFund,cash_flows_eff)).split(",")[2],axis=1)
            df['AssetType4']=df.apply(lambda r: (assetClass(r.Sec_type,r.Sec_code, r.Sec_name,r.TypeFund,cash_flows_eff)).split(",")[3],axis=1)
            df['AssetType5']=df.apply(lambda r: (assetClass(r.Sec_type,r.Sec_code, r.Sec_name,r.TypeFund,cash_flows_eff)).split(",")[4],axis=1)
            df['MarketValue']= np.where(df[['AssetType1']].isin(['B. Futures Exposure','Dividend Exposure']),0, df[['Market_price']])
            df['EffExposure']= df[['Market_price']]
            
            df['Close_price']=np.where((df['AssetType2'].isin(['Index Future']))&(df['Quantity'].values!=0),
                                                                  (df['Market_price'].values/df['Quantity'].values)/10, 
                                                                   df['Close_price'].values)
            # Futures insert
            if ~fut_flow.empty:
                fut_flow = pd.merge(fut_flow[['Port_code','Sec_code']], (df[['Trade_date','AssetType1','AssetType5','AssetType3','Sec_code','Sec_type','Close_price']]).drop_duplicates(['Sec_code']), on=['Sec_code'], how='left').fillna(0)
                fut_flow['Quantity']=0
                fut_flow['MarketValue']=0
                fut_flow['EffExposure']=0
                fut_flow['Trade_date']=startDate
                fut_flow=(fut_flow[['Trade_date','Port_code','AssetType1','AssetType5','AssetType3','Sec_code','Sec_type','Close_price','Quantity','MarketValue','EffExposure']]).copy()
                fut_flow=fut_flow[~(fut_flow.Sec_code=='NoFuture')]
               
            df=df[df.Port_code.isin(lst_fund)]
                                       
            dfprt=(df[['Trade_date','Port_code','AssetType1','AssetType5','AssetType3','Sec_code', 'Sec_type','Close_price','Quantity','MarketValue','EffExposure']]).copy()
            
            # Remove cash and other non-equity asset classes for multi-asset class
         #   dfprt['MarketValue']=np.where((dfprt.Port_code.map(lambda x:dic_om_index[x][2])=='M')&(dfprt.AssetType1!='Equity Exposure'), 0, dfprt.MarketValue.values)
         #   dfprt['EffExposure']=np.where((dfprt.Port_code.map(lambda x:dic_om_index[x][2])=='M')&(dfprt.AssetType1!='Equity Exposure'), 0, dfprt.EffExposure.values)
            dfprt= dfprt[~((dfprt.AssetType1=='Other')&(dfprt.Port_code.map(lambda x:dic_om_index[x][2])=='M'))]
            
            #Remove SSF Dividend Exposure
            dfprt.loc[:,'EffExposure']=np.where(dfprt[['AssetType5']].isin(override),0, dfprt[['EffExposure']])
            dfprt=dfprt[~(dfprt.Port_code.isnull())]
            dfprt=dfprt[~(dfprt.Quantity.isnull())]
            dfprt_preflow=dfprt.copy()
            
            # Add futures structureback               
                
            dfprt_preflow=dfprt_preflow.append(fut_flow,sort=True)  
              
            
                        
            if ~cash_flows_eff.empty:
                xx=cfvf(cash_flows_eff, newest_cash,startDate, lst_fund,bf=0.005)
                cash_flows_eff=cash_flows_eff.merge((xx[1])[['Port_code','Inflow_use']], on=['Port_code'], how='left')
                cash_flows_eff=cash_flows_eff[['Port_code', 'Inflow_use', 'Eff_cash', 'fut_sufx']]
                cash_flows_eff.columns=['Port_code', 'Inflow', 'Eff_cash', 'fut_sufx']
                cash_flows_eff['Trade_date']=startDate
                cash_flows_eff['AssetType1']='A. Total cash'
                cash_flows_eff['AssetType2']='Settled cash'
                cash_flows_eff['AssetType3']='Cash flow'
                cash_flows_eff['AssetType4']='Cash flow'
                cash_flows_eff['AssetType5']='Cash flow'
                cash_flows_eff['Sec_code']='ZAR'
                cash_flows_eff['Sec_type']='VAL'
                cash_flows_eff['Close_price']=1
                cash_flows_eff['Quantity']= cash_flows_eff[['Inflow']]
                cash_flows_eff['MarketValue']= cash_flows_eff[['Inflow']]
                cash_flows_eff['EffExposure']= cash_flows_eff[['Inflow']]
                
                cash_flows=cash_flows_eff[['Trade_date', 'Port_code','AssetType1', 'AssetType5', 'AssetType3', 'Sec_code','Sec_type','Close_price', 'Quantity','MarketValue','EffExposure']]
                
                chx_flw=xx[1]
            else:
                chx_flw=pd.DataFrame(columns=['Port_code','Inflow_use','ActFlow'])
            
            dfprt1=dfprt_preflow.append(cash_flows, sort=True)  
            dfprt=dfprt1.drop(['Sec_type'],axis=1)
            
            print("--- %s seconds ---" % (tm.time() - run_time))
            
            xdf= dfprt[dfprt.AssetType3.isin(['Cash on call','Val cash','Cash flow'])]
            xdf1=xdf
            xdf=xdf.groupby(['Trade_date','Port_code']).agg({'MarketValue':'sum'}).reset_index()
           # xdf.Port_code.unique().tolist()
            
#            cash_xls = pd.read_excel(newest_cash,sheet_name='Cash', 
#                                 converters={'Settle Date': pd.to_datetime, 'Trade date':pd.to_datetime,
#                                                    'Portfolio':str, 'Type':str, 
#                                                    'Security name':str,
#                                                    'Security Code':str,
#                                                    'Quantity':float,
#                                                    ' +/-':str,
#                                                    'Amount': float},
#                                )
            cash_xls = pd.read_excel(newest_cash,sheet_name='Detail', 
                         converters={'Portfolio':str,
                                     'Instrument':str,
                                     'Currency':str,
                                     'Type':str,
                                     'Sign':str,
                                     'Qty':float,
                                     'Amt Current': float,
                                     'Amt Gross': float,
                                     'Trade Date':pd.to_datetime,
                                     'Settle Date': pd.to_datetime 
                                            },
                                     )
            cash_xls.columns = [col.strip()  for col in cash_xls.columns]
            cash_xls=cash_xls.drop(['Currency','Amt Gross'],axis=1)
            cash_xls.columns = ['Portfolio','Security Code','Type',' +/-','Quantity','Amount','Trade date',	'Settle Date']

            
            cash_xls=cash_xls[cash_xls.Portfolio.isin(xdf.Port_code.unique().tolist())]
            cash_xls=cash_xls.copy()
            cash_xls.loc[:,'Value']=np.where(cash_xls[' +/-']=='-', -1.0*cash_xls['Amount'].values, cash_xls['Amount'].values)
            cc=cash_xls.groupby(['Settle Date','Portfolio']).agg({'Value':'sum'}).reset_index()
            
            #ff=pd.pivot_table(cc, values = 'Value', index=['Portfolio'], columns = 'Settle Date').reset_index()
            ff=pd.pivot_table(cc, values = 'Value', index=['Settle Date'], columns = 'Portfolio').reset_index()
            dtes=pd.bdate_range(start=cc['Settle Date'].min() +bdut.relativedelta(bdays=+1, holidays=hol.SouthAfrica()), end=cc['Settle Date'].min() +bdut.relativedelta(bdays=+3, holidays=hol.SouthAfrica()), holidays =hol.SouthAfrica())
            all_dta=(pd.merge(ff,pd.DataFrame(dtes.tolist(), columns=['Settle Date']), how='right', left_on=['Settle Date'], right_on=['Settle Date'] ).fillna(0))
            all_dta.iloc[:,1:all_dta.shape[1]]=all_dta.iloc[:,1:all_dta.shape[1]].cumsum()
            all_dta=all_dta.melt(id_vars=['Settle Date'], value_vars=all_dta.columns[1:],var_name='Portfolio', value_name='Cash')
            all_dta=all_dta[all_dta['Settle Date']==all_dta['Settle Date'].max()]
            all_dta=pd.merge(xdf, all_dta, how='left', left_on=['Port_code'], right_on=['Portfolio'])
            all_dta.loc[:,'Cash']=all_dta.Cash.fillna(0)
            all_dta.loc[:,'Settled_cash']=all_dta['Cash'].values+all_dta['MarketValue'].values
            all_dta=all_dta[['Port_code','Settled_cash']]
            
            return [all_dta,fut_sufx]
#        


"""    
'******************************************************************************************************************************************************************************    
'                                                                   Systematic rules to get trading action
'******************************************************************************************************************************************************************************    
"""


def CashFlowFlag(Eff_cash,Total_cash, mx_totcash, mn_totcash, mx_effcash, mn_effcash, md_totcash, flw, fut_exp):
    
    if Eff_cash > mx_effcash:
        if Total_cash >= mx_totcash:
            if Total_cash >= (mx_totcash+md_totcash):
                if flw==0:
                     return 'Trade Equity + Futures (midpoint)'
                else:
                     return 'Trade Equity + Futures (only flow)'
            elif Total_cash < (mx_totcash+md_totcash):
                return 'Trade Equity + Futures (midpoint)'
        elif Total_cash <= mn_totcash: # unlikely to occur
            if Total_cash <= (mn_totcash-md_totcash):
                if flw==0:
                     return 'Trade Equity + Futures (midpoint)'
                else:
                     return 'Trade Equity + Futures (only flow)'
            elif Total_cash > (mn_totcash-md_totcash):
                return 'Trade Equity + Futures (midpoint)'
            #return 'Trade Equity + Futures' # unlikely to occur
        else: 
            return 'Trade Futures only'
    elif Eff_cash < mn_effcash:
        if Total_cash <= mn_totcash:
             if Total_cash <= (mn_totcash-md_totcash):
                 if flw==0:
                     return 'Trade Equity + Futures (midpoint)'
                 else:
                     return 'Trade Equity + Futures (only flow)'
             elif Total_cash > (mn_totcash-md_totcash):
                return 'Trade Equity + Futures (midpoint)'
            #return 'Trade Equity + Futures'
        elif Total_cash >= mx_totcash:
            if Total_cash >= (mx_totcash+md_totcash):
                if flw==0:
                     return 'Trade Equity + Futures (midpoint)'
                else:
                     return 'Trade Equity + Futures (only flow)'
            elif Total_cash < (mx_totcash+md_totcash):
                return 'Trade Equity + Futures (midpoint)'
         #   return 'Trade Equity + Futures' # unlikely to occur
        else:
            return 'Trade Futures only'
    else:
        if Total_cash < mn_totcash:
           return 'Trade Equity'
        elif Total_cash > mx_totcash:
            if fut_exp > mx_totcash: # Need to think around how to deal with this event
                return 'No Action'
            else:
                return 'Trade Equity'
        else:
           return 'No Action'
    
"""    
'******************************************************************************************************************************************************************************    
'                                                                   Systematic rules to get trading targets
'******************************************************************************************************************************************************************************    
"""
    
      
def trade_calc(Flag, tgt_effcash, tgt_totcash, fut_code,  mx_effcash, mn_effcash, ovrd_effcash, aeff_cash, atot_cash, fnd_val, fut_price, fut_exp, flw):
    import numpy as np
    from write_excel import  chck_fut

#    if Flag == 'Trade Equity + Futures':
#        tgt_totcash = tgt_totcash
#        tgt_effcash = tgt_effcash
    if Flag == 'Trade Equity + Futures (only flow)':
        t_effcash=aeff_cash-flw
        t_totcash=atot_cash-flw
        if ((t_effcash>mx_effcash)or(t_effcash<mn_effcash)):
            t_effcash=tgt_effcash
    elif Flag == 'Trade Equity + Futures (midpoint)': 
        t_effcash=aeff_cash-flw
        t_totcash=tgt_totcash
        if ((t_effcash>mx_effcash)or(t_effcash<mn_effcash)):
            t_effcash=tgt_effcash
    elif Flag == 'Trade Futures only':
         x_trd = np.where(fut_code=="NoFuture", "No Trade",
                                    np.where((aeff_cash>mx_effcash), 'Buy', 
                                             np.where((aeff_cash<mn_effcash), 'Sell', 'No Trade')))
         
         x_trd2 = np.where((x_trd=="No Trade")&(~np.isnan(ovrd_effcash))&(not(fut_code=="NoFuture")),
                                    np.where((aeff_cash>tgt_effcash), 'Buy', 
                                             np.where((aeff_cash<tgt_effcash), 'Sell', 'No Trade')),x_trd)
         no_fut = np.where(np.isin(x_trd2, ['Buy','Sell']), np.rint(((aeff_cash-tgt_effcash)*fnd_val)/(fut_price*10)), 0)
         no_fut = chck_fut(no_fut, aeff_cash, mx_effcash, mn_effcash, tgt_effcash, fut_price, fnd_val)
         fut_exp1=((no_fut*10*fut_price)/fnd_val)+fut_exp
         if (atot_cash - fut_exp1) < 0:
              t_totcash = np.where(np.isnan(fut_exp), 0, fut_exp) + tgt_effcash
              t_effcash = tgt_effcash
         else:
             t_effcash = np.where((atot_cash - fut_exp1) < 0, (atot_cash - fut_exp), (atot_cash - fut_exp1))
             t_totcash = atot_cash
    
    elif Flag == 'Trade Equity':
        
         t_totcash = np.where(np.isnan(fut_exp), 0, fut_exp) + tgt_effcash
         t_effcash = tgt_effcash
     
    else:
        t_totcash = atot_cash
        t_effcash = aeff_cash
  

    return [np.round(t_effcash,20),np.round(t_totcash,20)]    


"""
'******************************************************************************************************************************************************************************    
'                                                                   Systematic rules to get trading targets (automatic)
'******************************************************************************************************************************************************************************    
"""
    
      
def trade_calc_automatic(p,Flag, tgt_effcash, tgt_totcash, fut_code,  mx_effcash, mn_effcash, ovrd_effcash, aeff_cash, atot_cash, fnd_val, fut_price, fut_exp, flw, curr_fut):
    import numpy as np
    from write_excel import  chck_fut

#    if Flag == 'Trade Equity + Futures':
#        tgt_totcash = tgt_totcash
#        tgt_effcash = tgt_effcash
    if Flag == 'Trade Equity + Futures (only flow)':
        t_effcash=aeff_cash-flw
        t_totcash=atot_cash-flw
        if ((t_effcash>mx_effcash)or(t_effcash<mn_effcash)):
            t_effcash=tgt_effcash
        no_fut = np.where(np.isnan(np.rint(((t_totcash-t_effcash-fut_exp)*fnd_val)/fut_price/10)),0,np.rint(((t_totcash-t_effcash-fut_exp)*fnd_val)/fut_price/10))
        
    elif Flag == 'Trade Equity + Futures (midpoint)': 
        t_effcash=aeff_cash-flw
        t_totcash=tgt_totcash
        if ((t_effcash>mx_effcash)or(t_effcash<mn_effcash)):
            t_effcash=tgt_effcash
        no_fut = np.where(np.isnan(np.rint(((t_totcash-t_effcash-fut_exp)*fnd_val)/fut_price/10)),0,np.rint(((t_totcash-t_effcash-fut_exp)*fnd_val)/fut_price/10))
        
    elif Flag == 'Trade Futures only':
         x_trd = np.where(fut_code=="NoFuture", "No Trade",
                                    np.where((aeff_cash>mx_effcash), 'Buy', 
                                             np.where((aeff_cash<mn_effcash), 'Sell', 'No Trade')))
         
         x_trd2 = np.where((x_trd=="No Trade")&(~np.isnan(ovrd_effcash))&(not(fut_code=="NoFuture")),
                                    np.where((aeff_cash>tgt_effcash), 'Buy', 
                                             np.where((aeff_cash<tgt_effcash), 'Sell', 'No Trade')),x_trd)
         no_fut = np.where(np.isin(x_trd2, ['Buy','Sell']), np.rint(((aeff_cash-tgt_effcash)*fnd_val)/(fut_price*10)), 0)
         no_fut = chck_fut(no_fut, aeff_cash, mx_effcash, mn_effcash, tgt_effcash, fut_price, fnd_val)
         fut_exp1=((no_fut*10*fut_price)/fnd_val)+fut_exp
         if (atot_cash - fut_exp1) < 0:
              t_totcash = np.where(np.isnan(fut_exp), 0, fut_exp) + tgt_effcash
              t_effcash = tgt_effcash
         else:
             t_effcash = np.where((atot_cash - fut_exp1) < 0, (atot_cash - fut_exp), (atot_cash - fut_exp1))
             t_totcash = atot_cash
    
    elif Flag == 'Trade Equity':
        
         t_totcash = np.where(np.isnan(fut_exp), 0, fut_exp) + tgt_effcash
         t_effcash = tgt_effcash
         no_fut = 0
     
    else:
        t_totcash = atot_cash
        t_effcash = aeff_cash
        no_fut = 0
    
    cash_bpm= (t_effcash*fnd_val)-((t_effcash-aeff_cash)*fnd_val)-(no_fut*10* np.nan_to_num(fut_price))
    eq_trd = cash_bpm - t_effcash*fnd_val
#    print(str(p+"=("+str(t_effcash)+"*"+str(fnd_val)+")-(("+str(t_effcash)+"-"+str(aeff_cash)+")*"+str(fnd_val)+")+(("+str(no_fut)+"*10*"+str( np.nan_to_num(fut_price))+"))"))
#    print(str(p+str(cash_bpm)+"eq trd: "+str(eq_trd)))
    
    tot_fut=np.nan_to_num(no_fut)+np.nan_to_num(curr_fut)
    
    return [np.round(t_effcash,20),np.round(t_totcash,20), no_fut, tot_fut, np.round(cash_bpm,8),np.round(eq_trd,8)]    
  
"""    


 
'******************************************************************************************************************************************************************************    
'                                                                   Bulk cash calc excel report
'******************************************************************************************************************************************************************************    
"""
def bulk_cash_excel_report(startDate,new_dat_pf,new_dat, n_comb,dic_users,dic_om_index,newest,output_folder,fnd_excp,chx_flw,automatic,dirtoimport_cashbal,orders,
                           vba_bin='//za.investment.int/dfs/dbshared/DFM/Tools/Indexation_trading_tools/IndexTrader/code/vbaProject.bin'):
    
    import pandas as pd
    import numpy as np
    import datetime as dt
    import os
    from datetime import datetime, timedelta
    import openpyxl as px
    from openpyxl.styles import colors, Font, Border, Side ,Protection
    #import openpyxl.celldown
    #from openpyxl import load_workbook
    from write_excel import select_fund as sf
    import xlsxwriter
    import time
    from write_excel import hedge_with_box as hf
    from write_excel import cash_forecast 
    
    start_time = datetime.now() 
    
    if automatic:
        auto_trade=True
    else:
        auto_trade=False
 #   output_folder= 'c:/data/'
    
    output_folder1=str(output_folder+'\\'+dic_users[os.environ.get("USERNAME").lower()][1])
    if not os.path.exists(output_folder1):
        os.makedirs(output_folder1)
        
    
    #print(output_folder1)
    output_file = output_folder1+'\\BatchCashCalc_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsx'
    st_row = 19
#    st_it = st_row+1
    
    st_col= 'C'
    #writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
    
    #hdr = new_dat_pf.columns.tolist()
    
    
    new = new_dat_pf.reset_index()
    new= new[~new.AssetType5.isin(['SSF DIV'])]
    
    new_flow = n_comb
    inv=(new_flow[['Port_code','fin_teff_cash','fin_tot_cash', 'InvType','CashFlowFlag','Min_EffCash','Max_EffCash','Min_TotalCash','Max_TotalCash']]).set_index('Port_code').T.to_dict('list')
    
    
    lst_fund = chx_flw.Port_code.unique()#sf(False)    
    h_w=hf(chx_flw,lst_fund)
    #new.pivot(index=['Port_code','AssetType1'], columns= 'AssetType3', values='EffExposure')
    
    #h=pd.pivot_table(new,  index=['Port_code'],columns=['AssetType1','AssetType3'], values=['EffExposure','EffWgt'],aggfunc=np.sum)
    for key,val in h_w.items():
  #      print(key)
  #      print(val)
        if val==1:
            print(inv[key][2])
            inv[key][2]='Hedged Withdrawal'
            print(inv[key][2])
    
    c_st=cash_forecast(dirtoimport_cashbal,startDate,lst_fund)
    
    
    
    qf=pd.DataFrame([], dtype='object')
    futures_dict = dict()
    cashflow_dict = dict()
    
    for fnd in lst_fund:
        fnd_sel= str(fnd+' Total')
        new_prt= new[new.Port_code.isin([fnd])]
        new_frt= new[~new.AssetType5.isin(new_prt.AssetType5.unique().tolist())]
        new_frt1=new_frt.copy()
        new_frt1.loc[:,'Port_code'] = fnd
        new_frt1.loc[:,'Close_price':'MktWgt'] = np.nan
        new_prt=new_prt.append(new_frt1)
     
        h=pd.pivot_table(new_prt,  index=['Port_code','AssetType1','AssetType5'],values=['EffExposure','EffWgt'], aggfunc=np.sum)
        g=pd.pivot_table(new_prt,  index=['Port_code','AssetType1','AssetType5'],values=['Quantity','Close_price'], aggfunc=np.sum)
        g=g[g.index.get_level_values(2).isin(['A. INDEX FUTURES'])]
        g.columns = ['EffExposure','EffWgt']
        g = g[['EffWgt','EffExposure']]
        g = g.reset_index()
        g['AssetType5'] = 'A. No. Futures'
        g.columns = ['Port_code', 'AssetType1', 'AssetType5', 'EffExposure', 'EffWgt']
        g=g.set_index(['Port_code','AssetType1','AssetType5'])
        
        # Create futures dictionary
        p=(new_prt[new_prt.AssetType5.isin(['A. INDEX FUTURES'])][['Port_code','AssetType3']]).drop_duplicates('Port_code').set_index('Port_code').T.to_dict('list')
        
        if fnd in fnd_excp:
            p[fnd] = ['No Future']
        
        # Cash flow dicitionary
        c=(((new_dat[(new_dat.index.get_level_values('AssetType5').isin(['Cash flow']))&(new_dat.index.get_level_values('Port_code').isin([fnd]))]['MarketValue']).reset_index())[['Port_code','MarketValue']]).set_index('Port_code').T.to_dict('list')
        
        # Inv Type dictionary
        
        
        futures_dict.update(p)
        cashflow_dict.update(c)
        
        
    #h.query('Port_code==["OMSI01"]')
    #h.stack()
    #h= h.T
    
        df1 = h.groupby(level=[0,1]).sum()
    
        df1.index = pd.MultiIndex.from_arrays([df1.index.get_level_values(0), 
                                               df1.index.get_level_values(1)+ '', 
                                               len(df1.index) * ['']])
        df2 = h.groupby(level=0).sum() 
        df2['EffExposure'] = df2.EffExposure.values-(df1[(df1.index.get_level_values(1).isin(['B. Futures Exposure']))].reset_index()).EffExposure.values
        df2['EffWgt'] = 1
        
        df2.index = pd.MultiIndex.from_arrays([df2.index.values + ' Total',
                                               len(df2.index) * [''], 
                                               len(df2.index) * ['']])
        df2_1=df2.reset_index()
        df2_1.loc[df2_1.level_0.isin([fnd_sel]),'EffExposure'] = new_prt.FundValue.head(1).values
        df2_1.loc[df2_1.level_0.isin([fnd_sel]),'EffWgt'] = 1
        df2 = df2_1.set_index(['level_0','level_1','level_2'])
    
    #df2.index.get_level_values('CORPEQ Total').isin(['CORPEQ'])
    #df = pd.concat([h, df1, df2]).sort_index(level=[0])
        df = pd.concat([df1,df2,h,g]).sort_index(level=[1])
        df=df.reset_index(level=0, drop=True)
        d = dict(zip(df.columns, ['EffExp', 'EffWgt']))
        df=df.rename(columns=d,level=0)
       # df['EffExposure']=pd.Series(["R{: ,.2f}".format(val) for val in df['EffExposure']], index = df.index)
       # df['EffWgt']=pd.Series(["{0:.4f}%".format(val * 100) for val in df['EffWgt']], index = df.index)
    #wf =pd.concat([df1,df2,h]).sort_index(level=[1])
    #df = pd.concat([h, df2]).sort_index(level=[0])
    
        qf= pd.concat([qf,df], axis=1)
        
    
    qf=qf.rename(index={'':'Fund Value','A. Total cash':'Total cash','B. Futures Exposure':'Futures Exposure'}, level=0)    
    qf=qf.rename(index={'A. VAL':'VAL', 
                        'B. DIF':'DIF',
                        'A. INDEX FUTURES':'INDEX FUTURES',
                        'A. No. Futures':'No. Futures',
                        'B. SSF':'SSF',
                        'C. CALL': 'CALL',
                        'D. SAFEX':'SAFEX'}, level=1)    
    
        
    #h.to_excel(writer, sheet_name='Sheet1', startrow=st_row, header=  hdr,index_label = ['Portfolio Code','Fund', 'Level 1','Level 2', 'Level 3'])
    
        
        
    
    writer = pd.ExcelWriter(output_file, engine='xlsxwriter')
        
    #wf.to_excel(writer, sheet_name='Sheet1', startrow=st_row, startcol =5)
    
    qf.to_excel(writer, sheet_name='Summary', startrow=st_row, startcol =0, header = False)
    
    workbook  = writer.book
    
    workbook.filename= output_folder1+'\\BatchCashCalc_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsm'
    #workbook.add_vba_project('C:/IndexTrader/code/vbaProject.bin')
    workbook.add_vba_project(vba_bin)
    
    
    #writer.save()
    
    
    worksheet = writer.sheets['Summary']
    
    
    cell_format1 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':12})
    cell_format2 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':11})
    cell_format2_1 = workbook.add_format({'bold': False, 'font_color': 'black', 'font':11})
    cell_format2_2 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':11,  'border':1})
    cell_format2_2.set_text_wrap() 
    cell_format2_2.set_font_name('Calibri')
    cell_format2_3 = workbook.add_format({'bold': False, 'font_color': 'black', 'font':11,'bg_color':'#CCFFFF'})
    cell_format2_3.set_font_size(10)
    cell_format2_4 = workbook.add_format({'bold': True, 'font_color': 'white', 'font':11,'bg_color':'#003366'})
    cell_format2_5 = workbook.add_format({'bold': False, 'font_color': 'black', 'font':10})                                         
    cell_format2_5.set_font_size(10)
    cell_format2_5.set_align('right')
    
    cell_format3 = workbook.add_format({'bold': True, 'bg_color':'#CCFFFF', 'font':11, 'locked':False })
    #cell_format4 = workbook.add_format({'bold': True, 'bg_color':'#339966', 'font':11, 'locked':False })
    cell_format5 = workbook.add_format({'bold': True, 'bg_color':'#C0C0C0', 'font':11, 'align': 'center','border': 1})
    cell_format6 = workbook.add_format({'bold': True, 'bg_color':'#CCFFFF', 'font':11,  'align': 'center','border':1})
    
    cell_format5_1 = workbook.add_format({'bold': False, 'bg_color':'#C0C0C0', 'font':10, 'locked':False,'num_format': 'R#,##0','font_color': 'red'  })
    cell_format5_2 = workbook.add_format({'bold': False, 'bg_color':'#C0C0C0', 'font':10, 'locked':False,'num_format': '0.00%','font_color': 'red'  })
    
    cell_format6_1 = workbook.add_format({'bold': False, 'bg_color':'#CCFFFF', 'font':10, 'locked':False,'num_format': 'R#,##0', 'font_color': 'red' })
    cell_format6_2 = workbook.add_format({'bold': False, 'bg_color':'#CCFFFF', 'font':10, 'locked':False,'num_format': '0.00%', 'font_color': 'red' })
    
    cell_format5_3 = workbook.add_format({'bold': False, 'bg_color':'#C0C0C0', 'font':10, 'locked':False,'num_format': 'R#,##0','font_color': 'black'  })
    cell_format5_4 = workbook.add_format({'bold': False, 'bg_color':'#C0C0C0', 'font':10, 'locked':False,'num_format': '0.00%','font_color': 'black'  })
    
    cell_format6_3 = workbook.add_format({'bold': False, 'bg_color':'#CCFFFF', 'font':10, 'locked':False,'num_format': 'R#,##0', 'font_color': 'black' })
    cell_format6_4 = workbook.add_format({'bold': False, 'bg_color':'#CCFFFF', 'font':10, 'locked':False,'num_format': '0.00%', 'font_color': 'black' })
    
    cell_format6_5 = workbook.add_format({'bold': True, 'bg_color':'#FFFFCC',   'font_color': 'black' ,'align': 'center','border': 1})
    cell_format6_6 = workbook.add_format({'bold': False, 'font_color': 'blue', 'locked': False ,'align': 'center','border': 1})
    cell_format6_6.set_font_size(10)
    cell_format6_7 = workbook.add_format({'bold': True, 'bg_color':'#00CCFF',   'font_color': 'black' ,'align': 'center','border': 1})
    
    cell_format7 = workbook.add_format({'bold': True, 'bg_color':'#CCFFCC', 'font':11, 'align': 'center','border': 1})
    cell_format8 = workbook.add_format({'bold': True, 'font':11, 'font_color': '#339966','align': 'center','border': 1})
    cell_format9 = workbook.add_format({'bold': True, 'font':11, 'bg_color': '#FFFF00','align': 'center','border': 0})
    
                                        
    td_cell_format_1 = workbook.add_format({'bold': True, 'bg_color':'#FFC7CE', 'font':10,'font_color': '#9C0006'  })
    td_cell_format_2 = workbook.add_format({'bold': True, 'bg_color':'#C6EFCE', 'font':10,'font_color': '#006100'  })
    
    #td_cell_format_3 = workbook.add_format({'bold': True, 'bg_color':'#FF0000', 'font':10, 'locked':False,'font_color': '#9C0006','align': 'center'  })
    td_cell_format_4 = workbook.add_format({'bold': True, 'bg_color':'#CCCCFF', 'font':11, 'locked':False,'font_color': '#0000FF','align': 'center'  })
    td_cell_format_4.set_font_size(11)
                                          
    format1 = workbook.add_format({'num_format': 'R#,##0'})
    format1_1 = workbook.add_format({'num_format': 'R#,##0', 'locked': False})
    format2 = workbook.add_format({'num_format': '0.000%'})     
    
    
    format1_b = workbook.add_format({'bold': True,'num_format': 'R#,##0'})
    format2_b = workbook.add_format({'bold': True,'num_format': '0.000%'})                                    
    
    worksheet.write_string('A1', 'Batch Fund Cash Calc',cell_format1)
      
    #        worksheet.write('A2', '<portfolio owner>')
    worksheet.write('A3', 'Date', cell_format2)
    #worksheet.write('B1', fund)
    #worksheet.write('B2', Manager)
    worksheet.write('B3', datetime.strftime(datetime.today(), "%Y-%m-%d %H:%M:%S"), cell_format2_1)
    worksheet.write_string('A4', 'Prepared by',cell_format2)
    worksheet.write_string('B4',str(dic_users[os.environ.get("USERNAME").lower()][1]).upper(), cell_format2_1)
    worksheet.write_string('A6', 'Authorised by',cell_format2)
    worksheet.merge_range('B6:C6','', cell_format3)
    worksheet.write_string('A5', 'Orders included',cell_format2)
    worksheet.write_string('B5',str(orders), cell_format9)
    
    
    worksheet.write_string('A10', 'CASHFLOWS',cell_format1 )
    worksheet.write_string('B10', 'Type of Flow',cell_format1 )
    worksheet.write_string('B11', 'Amount Flow',cell_format1 )
    
    worksheet.write_string('A13', 'TARGET CASH LEVELS',cell_format1 )
    worksheet.write_string('B14', 'Target Cash Exposure',cell_format1 )
    worksheet.write_string('B15', 'Target Total Cash',cell_format1 )
    worksheet.write_string('B16', 'Futures',cell_format1 )
    worksheet.write_string('B17', 'Benchmark',cell_format1 )
    
    worksheet.write_string('D3', 'File used:',cell_format2)
    worksheet.write_string('E3', newest,cell_format2_1)
    worksheet.write_string('D4', 'Timestamp:',cell_format2)
    worksheet.write_string('E4', time.ctime(os.path.getmtime(newest)) ,cell_format2_1)
   
    
    
    
    #import string  
    #alp=list(string.ascii_uppercase) 
    alp=[]
    def colnum_string(n):
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string
        #print(string)
    #print (colnum_string(1))
    
    
    for a in range(1,(len(lst_fund))*2+5):
    #    print(a)
        x_alp=colnum_string(a)
    #    print(x_alp)
        alp.append(x_alp)
    #    print(alp)
        
    
    
    st_col=2
    fmts=[cell_format5,cell_format6]
    fmts2_1=[cell_format5_1,cell_format6_1] #red R
    fmts2_2=[cell_format5_3,cell_format6_3] # black R
    
    fmts3_1=[cell_format5_2,cell_format6_2] # red %
    fmts3_2=[cell_format5_4,cell_format6_4] # black %
    cell_format1_1 = workbook.add_format({'bold': True, 'font_color': 'black', 'font':12, 'locked': False})
    
    jet=0
    worksheet.set_column('A:A', 24)
    worksheet.set_column('B:B', 24)
    
    for j in lst_fund:
     #   print(j)
        get_pl=str(alp[st_col]+str(st_row-12)+":"+alp[st_col+1]+str(st_row-12))
        get_pl2=str(alp[st_col]+str(st_row-8))
        get_pl3=str(alp[st_col+1]+str(st_row-8))
        get_pl4=str(alp[st_col]+str(st_row-9))
        get_pl5=str(alp[st_col]+str(st_row-5))
        get_pl6=str(alp[st_col]+str(st_row-4))
        get_pl7=str(alp[st_col+1]+str(st_row-5))
        get_pl8=str(alp[st_col+1]+str(st_row-4))
    
        worksheet.set_column(str(alp[st_col]+":"+alp[st_col]), 15,format1)
        worksheet.set_column(str(alp[st_col+1]+":"+alp[st_col+1]), 11,format2)
        worksheet.merge_range(get_pl,j, fmts[jet])
        worksheet.write_number(get_pl2, 0,fmts2_2[jet])
        worksheet.write_string(str(alp[st_col]+str(st_row-6)), 'Target',cell_format2)
        worksheet.write_string(str(alp[st_col+1]+str(st_row-6)), 'Post Trade',cell_format2)
        #worksheet.write_formula(get_pl3,str('='+get_pl2+'/('+str(alp[st_col]+str(st_row+1))+'+'+get_pl2+')'),fmts[jet])
        worksheet.write_formula(get_pl3,str('='+get_pl2+'/('+str(alp[st_col]+str(st_row+16))+')'),fmts[jet])
        worksheet.write_number(str(alp[st_col]+str(st_row-8)), cashflow_dict[j][0] ,format1_1)
        
        worksheet.write_string(get_pl4, inv[j][2] ,cell_format1_1)
        
        worksheet.data_validation(get_pl4, {'validate': 'list',
                                            'source': ['Investment', 
                                                       'Hedged Withdrawal', 
                                                 #      'Hedged With Pay(t)',
                                                       'Withdrawal Pay(t)',
                                                       'No cash flow'],
                                            'input_title': 'Enter a cash flow',
                                            'input_message': 'Type & Value'} )
        
        if auto_trade:
            worksheet.write_number(str(alp[st_col]+str(st_row-5)), inv[j][0] ,fmts3_2[jet])
            worksheet.write_number(str(alp[st_col]+str(st_row-4)), inv[j][1] ,fmts3_2[jet])
            worksheet.conditional_format(get_pl5,{'type': 'cell','criteria': '!=','value': inv[j][0] ,'format': td_cell_format_4})
            worksheet.conditional_format(get_pl6,{'type': 'cell','criteria': '!=','value': inv[j][1] ,'format': td_cell_format_4})
        
        else: 
            worksheet.write_formula(get_pl5, str('='+str(alp[st_col+1]+str(st_row+26))),fmts3_2[jet])
            worksheet.write_formula(get_pl6, str('='+str(alp[st_col+1]+str(st_row+17))),fmts3_2[jet])
            worksheet.conditional_format(get_pl5,{'type': 'cell','criteria': '!=','value': str('='+str(alp[st_col+1]+str(st_row+26))) ,'format': td_cell_format_4})
            worksheet.conditional_format(get_pl6,{'type': 'cell','criteria': '!=','value': str('='+str(alp[st_col+1]+str(st_row+17))) ,'format': td_cell_format_4})
        
        
        worksheet.write_formula(get_pl7, str('='+str(alp[st_col+1]+str(st_row+56))),fmts3_2[jet])
        worksheet.write_formula(get_pl8, str('='+str(alp[st_col+1]+str(st_row+47))),fmts3_2[jet])
        
        worksheet.conditional_format(get_pl2, {'type': 'cell','criteria': '<','value': 0,'format': fmts2_1[jet]})                              
        worksheet.conditional_format(get_pl2, {'type': 'cell','criteria': '>=','value': 0,'format': fmts2_2[jet]})   
        worksheet.conditional_format(get_pl3, {'type': 'cell','criteria': '<','value': 0,'format': fmts3_1[jet]})                              
        worksheet.conditional_format(get_pl3, {'type': 'cell','criteria': '>=','value': 0,'format': fmts3_2[jet]})      
        
        worksheet.conditional_format(get_pl4,{'type': 'cell','criteria': 'equal to','value': '"No cash flow"','format': cell_format1_1})
        worksheet.conditional_format(get_pl4,{'type': 'cell','criteria': '!=','value': '"No cash flow"','format': td_cell_format_4})
        
        worksheet.merge_range(str(alp[st_col]+str(st_row-3)+":"+alp[st_col+1]+str(st_row-3)), futures_dict[j][0] ,cell_format7)
        worksheet.merge_range(str(alp[st_col]+str(st_row-2)+":"+alp[st_col+1]+str(st_row-2)), dic_om_index[j][1] ,cell_format8)
       
        
        for g in range(st_row, st_row+ 25):
            if g in [st_row+1, st_row+2,st_row+7, st_row+11,st_row+13]:
                
                worksheet.conditional_format(str(alp[st_col]+str(g)),{'type': 'cell','criteria': '!=','value': inv[j][0] ,'format': format1_b})
                worksheet.conditional_format(str(alp[st_col+1]+str(g)),{'type': 'cell','criteria': '!=','value': inv[j][0] ,'format': format2_b})
        
        #worksheet.conditional_format(get_pl4,{'type': 'cell','criteria': 'equal to','value': '"No cash flow"','format': cell_format1})
        
        st_col=st_col+2
        jet=np.where(jet==0,1,0)
       # print(get_pl)
    
    # grouping categories
    del g
    
    worksheet.set_row((st_row+2), None, None, {'level': 1})
    worksheet.set_row((st_row+3), None, None, {'level': 1})
    worksheet.set_row((st_row+4), None, None, {'level': 1})
    worksheet.set_row((st_row+5), None, None, {'level': 1})
    #worksheet.set_row((st_row+6), None, None, {'level': 1})
    
    worksheet.set_row((st_row+7), None, None, {'level': 1})
    worksheet.set_row((st_row+8), None, cell_format2_3,{'level': 1})
    worksheet.set_row((st_row+9), None, None, {'level': 1})
    
    worksheet.set_row((st_row+11), None, None, {'level': 1})
    worksheet.set_row((st_row+13), None, None, {'level': 1})
    
    worksheet.set_row(st_row-1, None, cell_format2_4)
    worksheet.merge_range(str('A'+str(st_row)+':B'+str(st_row)),'PRE-TRADE', cell_format2_4)
    
    # end grouping
    def form_at(st_no=(st_row+len(qf)),_label_='INCLUDING FLOWS'):
        in_st=st_no
        in_cell_format1 = workbook.add_format({'bold': True, 'font_color': 'black', 'border':1, 'align': 'center', 'valign': 'top' })
        bn_cell_format1 = workbook.add_format({'bold': True, 'font_color': 'black', 'border':1, 'align': 'center', 'valign': 'top','bg_color':'#FF8080' })
        
        worksheet.set_row(in_st, None, cell_format2_4)
        worksheet.merge_range(str('A'+str(in_st+1)+':B'+str(in_st+1)),_label_, cell_format2_4)
        worksheet.write_string(str('A'+str(in_st+2)), 'Fund Value',in_cell_format1)
        worksheet.merge_range(str('A'+str(in_st+3)+':A'+str(in_st+7)),'Total cash', in_cell_format1)
        worksheet.merge_range(str('A'+str(in_st+8)+':A'+str(in_st+11)),'Futures Exposure', in_cell_format1)
        worksheet.merge_range(str('A'+str(in_st+12)+':A'+str(in_st+13)),'Effective cash', in_cell_format1)
        worksheet.merge_range(str('A'+str(in_st+14)+':A'+str(in_st+15)),'Equity Exposure', in_cell_format1)
        
        worksheet.write_string(str('B'+str(in_st+2)), '',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+3)), '',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+4)), 'VAL',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+5)), 'DIF',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+6)), 'CALL',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+7)), 'SAFEX',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+8)), '',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+9)), 'INDEX FUTURES',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+10)), 'No. Futures',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+11)), 'SSF',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+12)), '',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+13)), 'null',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+14)), '',in_cell_format1)
        worksheet.write_string(str('B'+str(in_st+15)), 'EQUITY',in_cell_format1)
        
        worksheet.set_row((in_st+3), None, None, {'level': 1})
        worksheet.set_row((in_st+4), None, None, {'level': 1})
        worksheet.set_row((in_st+5), None, None, {'level': 1})
        worksheet.set_row((in_st+6), None, None, {'level': 1})
        #worksheet.set_row((st_row+6), None, None, {'level': 1})
        
        worksheet.set_row((in_st+8), None, None, {'level': 1})
        worksheet.set_row((in_st+9), None, cell_format2_3,{'level': 1})
        worksheet.set_row((in_st+10), None,None,  {'level': 1})
        
        worksheet.set_row((in_st+12), None, None, {'level': 1})
        worksheet.set_row((in_st+14), None, None, {'level': 1})
        
        if _label_=='POST-TRADE':
             worksheet.write_string(str('B'+str(in_st+17)), 'BARRA CASH',bn_cell_format1)
             worksheet.write_string(str('B'+str(in_st+18)), 'TRADE ACTION',cell_format6_5)
             worksheet.write_string(str('B'+str(in_st+20)), 'CASH BALANCES',cell_format6_7)
             worksheet.write(str('B'+str(in_st+21)), datetime.strftime(c_st[1][0], "%Y-%m-%d"), cell_format2_5)
             worksheet.write(str('B'+str(in_st+22)), datetime.strftime(c_st[1][1], "%Y-%m-%d"), cell_format2_5)
             worksheet.write(str('B'+str(in_st+23)), datetime.strftime(c_st[1][2], "%Y-%m-%d"), cell_format2_5)
             
    
    form_at()
    form_at(st_no=(st_row+2*len(qf)+1),_label_='TRADE')
    form_at(st_no=(st_row+3*len(qf)+2),_label_='POST-TRADE')
    
    
    # Write formula
    
    # Inflow
    
    in_st = (st_row+1*len(qf)+1)
    in_col=2
    
    informat1_n = workbook.add_format({'bold': False,'num_format': 'R#,##0'})
    informat2_n = workbook.add_format({'bold': False,'num_format': '0.000%'})                                    
    informat1_b = workbook.add_format({'bold': True,'num_format': 'R#,##0'})
    informat2_b = workbook.add_format({'bold': True,'num_format': '0.000%'})                   
    
                
    for f in lst_fund:
      #  print(f)
        worksheet.write_formula(str(alp[in_col]+str(in_st+1)),str('='+str(alp[in_col]+str(in_st+2))+'+'+str(alp[in_col]+str(in_st+13))),informat1_b)
        worksheet.write_formula(str(alp[in_col]+str(in_st+2)),str('=SUM('+str(alp[in_col]+str(in_st+3))+':'+str(alp[in_col]+str(in_st+6))+')'),informat1_b)
        worksheet.write_formula(str(alp[in_col]+str(in_st+3)),str('='+str(alp[in_col]+str(in_st-12))+'+IF('+str(alp[in_col]+str(in_st-24))+\
                                    '="Hedged Withdrawal",0,'+str(alp[in_col]+str(in_st-23))+')'),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+4)),str('='+str(alp[in_col]+str(in_st-11))),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+5)),str('='+str(alp[in_col]+str(in_st-10))),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+6)),str('='+str(alp[in_col]+str(in_st-9))),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+7)),str('=SUM('+str(alp[in_col]+str(in_st+8))+','+str(alp[in_col]+str(in_st+10))+')'),informat1_b)
        worksheet.write_formula(str(alp[in_col]+str(in_st+8)),str('='+str(alp[in_col]+str(in_st-7))),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+10)),str('='+str(alp[in_col]+str(in_st-5))),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+11)),str('='+str(alp[in_col]+str(in_st+12))),informat1_b)
        worksheet.write_formula(str(alp[in_col]+str(in_st+12)),str('='+str(alp[in_col]+str(in_st+2))+'-'+str(alp[in_col]+str(in_st+7))),informat1_n)
        worksheet.write_formula(str(alp[in_col]+str(in_st+13)),str('='+str(alp[in_col]+str(in_st+14))),informat1_b)
        worksheet.write_formula(str(alp[in_col]+str(in_st+14)),str('='+str(alp[in_col]+str(in_st-1))),informat1_n)
        
        for g in range(in_st+1, in_st+qf.shape[0]+1):
            if g in [in_st+1,in_st+2,in_st+7,in_st+11,in_st+13]:
                worksheet.write_formula(str(alp[in_col+1]+str(g)),str('='+str(alp[in_col]+str(g))+'/'+str(alp[in_col]+'$'+str(in_st+1))),informat2_b)
            #    print(str('B='+str(alp[in_col]+str(g))+'/'+str(alp[in_col]+'$'+str(in_st+1))))
            elif g!=(in_st+9):
                worksheet.write_formula(str(alp[in_col+1]+str(g)),str('='+str(alp[in_col]+str(g))+'/'+str(alp[in_col]+'$'+str(in_st+1))),informat2_n)
            #    print(str('N='+str(alp[in_col]+str(g))+'/'+str(alp[in_col]+'$'+str(in_st+1))))
            else:
            #    print(g)
                gg=1
        in_col=in_col+2
    
    del f 
    del g
    del gg
    # Trade
    
    td_st = (st_row+2*len(qf)+2)
    td_col=2
    
    
    for f in lst_fund:
        worksheet.write_formula(str(alp[td_col]+str(td_st+1)),str('='+str(alp[td_col]+str(td_st+8))+'+'+str(alp[td_col]+str(td_st+14))),informat1_b)
        worksheet.write_formula(str(alp[td_col]+str(td_st+2)),str('=SUM('+str(alp[td_col]+str(td_st+3))+':'+str(alp[td_col]+str(td_st+6))+')'),informat1_b)
        worksheet.write_formula(str(alp[td_col]+str(td_st+3)),str('=-'+str(alp[td_col]+str(td_st+6))),informat1_n)
        worksheet.write_formula(str(alp[td_col]+str(td_st+4)),str('='+str(alp[td_col]+str(td_st+8))+'+'+str(alp[td_col]+str(td_st+11))),informat1_n)
        worksheet.write_formula(str(alp[td_col]+str(td_st+6)),\
                                str('=IF(ISNUMBER('+str(alp[td_col]+str(td_st+9))+'*('+str(alp[td_col]+str(td_st-24))+'/'+ \
                                                        str(alp[td_col]+str(td_st-21))+')),'+str(alp[td_col]+str(td_st+9))+ \
                                                        '*('+str(alp[td_col]+str(td_st-24))+'/'+str(alp[td_col]+str(td_st-21))+'),0)'), informat1_n)
         
        worksheet.write_formula(str(alp[td_col]+str(td_st+7)),str('='+str(alp[td_col]+str(td_st+8))),informat1_b)
        worksheet.write_formula(str(alp[td_col]+str(td_st+8)),\
                                str('=IF(ISERROR('+str(alp[td_col]+str(td_st+9))+'*'+str(alp[td_col+1]+str(td_st-21))+'*10),0,'+\
                                                       str(alp[td_col]+str(td_st+9))+'*'+str(alp[td_col+1]+str(td_st-21))+'*10)'), informat1_n)
        worksheet.write_formula(str(alp[td_col]+str(td_st+9)),\
                  #              str('=IF('+str(alp[td_col]+str(td_st-33))+'="No Future",0,INT((ROUND('+str(alp[td_col]+str(td_st-34))+'-'+str(alp[td_col]+str(td_st-35))+\
                                str('=IF('+str(alp[td_col]+str(td_st-33))+'="No Future",0,IF(('+str(alp[td_col]+str(td_st-34))+'-(IF('+str(alp[td_col]+str(st_row-9))+'="Hedged Withdrawal",'+str(alp[td_col+1]+str(st_row-8))+',0))'+'-'+str(alp[td_col]+str(td_st-35))+\
                                    '-'+str(alp[td_col+1]+str(td_st-8))+')>0,ROUNDDOWN((('+str(alp[td_col]+str(td_st-34))+'-(IF('+str(alp[td_col]+str(st_row-9))+'="Hedged Withdrawal",'+str(alp[td_col+1]+str(st_row-8))+',0))'+'-'+str(alp[td_col]+str(td_st-35))+\
                                    '-'+str(alp[td_col+1]+str(td_st-8))+')*'+str(alp[td_col]+str(td_st-14))+')/'+ \
                                        str(alp[td_col+1]+str(td_st-21))+'/10,0),ROUNDUP(ROUND((('+str(alp[td_col]+str(td_st-34))+'-(IF('+str(alp[td_col]+str(st_row-9))+'="Hedged Withdrawal",'+str(alp[td_col+1]+str(st_row-8))+',0))'+'-'+str(alp[td_col]+str(td_st-35))+\
                                    '-'+str(alp[td_col+1]+str(td_st-8))+')*'+str(alp[td_col]+str(td_st-14))+')/'+ \
                                        str(alp[td_col+1]+str(td_st-21))+'/10,9),0)'+'))'),cell_format2_1)
        worksheet.write_formula(str(alp[td_col+1]+str(td_st+9)),str('=IF('+str(alp[td_col]+str(td_st+9))+'<0,"SOC",IF('+str(alp[td_col]+str(td_st+9))\
                                                                +'>0,"BOC",""))'),informat1_n) 
        
        worksheet.write_formula(str(alp[td_col]+str(td_st+11)),str('='+str(alp[td_col]+str(td_st+12))),informat1_n)
        worksheet.write_formula(str(alp[td_col]+str(td_st+12)),str('=('+str(alp[td_col]+str(td_st-14))+'*'+str(alp[td_col]+str(td_st-35))+')-'\
                                    +str(alp[td_col]+str(td_st-4))),informat1_n)
        worksheet.write_formula(str(alp[td_col]+str(td_st+13)),str('='+str(alp[td_col]+str(td_st+14))),informat1_n)
        worksheet.write_formula(str(alp[td_col]+str(td_st+14)),str('=-'+str(alp[td_col]+str(td_st+4))),informat1_n)
        worksheet.write_formula(str(alp[td_col+1]+str(td_st+14)),str('=IF('+str(alp[td_col]+str(td_st+14))+'<-0.01,"SELL EQTY",IF('+str(alp[td_col]+str(td_st+14))\
                                                                +'>0.01,"BUY EQTY",""))'),informat1_n) 
        
        for g in range(td_st+1, td_st+qf.shape[0]+1):
          #  print(g)
            if g in [td_st+1,td_st+2,td_st+7,td_st+11,td_st+13]:
                worksheet.write_formula(str(alp[td_col+1]+str(g)),str('='+str(alp[td_col]+str(g))+'/'+str(alp[td_col]+'$'+str(td_st-14))),informat2_b)
         #       print(str('B='+str(alp[td_col]+str(g))+'/'+str(alp[td_col]+'$'+str(td_st+1))))
            elif g in [td_st+3,td_st+4,td_st+6,td_st+8,td_st+12]:
                worksheet.write_formula(str(alp[td_col+1]+str(g)),str('='+str(alp[td_col]+str(g))+'/'+str(alp[td_col]+'$'+str(td_st-14))),informat2_n)
         #       print(str('N='+str(alp[td_col]+str(g))+'/'+str(alp[td_col]+'$'+str(td_st+1))))
            else:
                gg=1
            
            worksheet.conditional_format(str(alp[td_col]+str(td_st+9)), {'type': 'cell','criteria': '<','value': 0,'format': td_cell_format_1})    
            worksheet.conditional_format(str(alp[td_col]+str(td_st+9)), {'type': 'cell','criteria': '>','value': 0,'format': td_cell_format_2})
            worksheet.conditional_format(str(alp[td_col+1]+str(td_st+9)),{'type': 'cell','criteria': '=','value': '"SOC"','format': td_cell_format_1}) 
            worksheet.conditional_format(str(alp[td_col+1]+str(td_st+9)),{'type': 'cell','criteria': '=','value': '"BOC"','format': td_cell_format_2}) 
            worksheet.conditional_format(str(alp[td_col+1]+str(td_st+14)),{'type': 'cell','criteria': '=','value': '"SELL EQTY"','format': td_cell_format_1}) 
            worksheet.conditional_format(str(alp[td_col+1]+str(td_st+14)),{'type': 'cell','criteria': '=','value': '"BUY EQTY"','format': td_cell_format_2}) 
            worksheet.conditional_format(str(alp[td_col]+str(td_st+14)),{'type': 'cell','criteria': '<','value': -0.01,'format': td_cell_format_1})
            worksheet.conditional_format(str(alp[td_col]+str(td_st+14)),{'type': 'cell','criteria': '>','value': 0.01,'format': td_cell_format_2}) 
        
        td_col=td_col+2
                                 
    del f 
    del g
    del gg
    
    # Post trade
    
    pt_st = (st_row+3*len(qf)+3)
    pt_col=2
    pt_cell_format_1 = workbook.add_format({'bold': True, 'bg_color':'#FFC7CE', 'font':10, 'locked':False,'font_color': '#9C0006','num_format': '0.000%' })
    pt_cell_format_2 = workbook.add_format({'bold': True, 'bg_color':'#C6EFCE', 'font':10, 'locked':False,'font_color': '#006100','num_format': '0.000%'  })
    
    for f in lst_fund:
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+1)),str('='+str(alp[pt_col]+str(pt_st+2))+'+'+str(alp[pt_col]+str(pt_st+13))),informat1_b)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+2)),str('=SUM('+str(alp[pt_col]+str(pt_st+3))+':'+str(alp[pt_col]+str(pt_st+6))+')'),informat1_b)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+3)),str('='+str(alp[pt_col]+str(pt_st-27))+'+'+str(alp[pt_col]+str(pt_st-12))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+4)),str('='+str(alp[pt_col]+str(pt_st-26))+'+'+str(alp[pt_col]+str(pt_st-11))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+5)),str('='+str(alp[pt_col]+str(pt_st-25))+'+'+str(alp[pt_col]+str(pt_st-10))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+6)),str('='+str(alp[pt_col]+str(pt_st-24))+'+'+str(alp[pt_col]+str(pt_st-9))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+7)),str('='+str(alp[pt_col]+str(pt_st-23))+'+'+str(alp[pt_col]+str(pt_st-8))),informat1_b)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+8)),str('='+str(alp[pt_col]+str(pt_st-22))+'+'+str(alp[pt_col]+str(pt_st-7))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+9)),str('='+str(alp[pt_col]+str(pt_st-36))+'+'+str(alp[pt_col]+str(pt_st-6))),cell_format2_3)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+10)),str('='+str(alp[pt_col]+str(pt_st-35))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+11)),str('='+str(alp[pt_col]+str(pt_st+12))),informat1_b)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+12)),str('='+str(alp[pt_col]+str(pt_st-18))+'+'+str(alp[pt_col]+str(pt_st-3))),informat1_n)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+13)),str('='+str(alp[pt_col]+str(pt_st+14))),informat1_b)
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+14)),str('='+str(alp[pt_col]+str(pt_st-16))+'+'+str(alp[pt_col]+str(pt_st-1))),informat1_n)
         
         worksheet.write_formula(str(alp[pt_col]+str(pt_st+16)),str('='+str(alp[pt_col]+str(pt_st+11))+'-'+str(alp[pt_col]+str(pt_st-13))),informat1_n)
         
         
         for g in range(pt_st+1, pt_st+qf.shape[0]+3):
         #   print(g)
            if g in [pt_st+1,pt_st+2,pt_st+7,pt_st+11,pt_st+13,pt_st+16]:
                worksheet.write_formula(str(alp[pt_col+1]+str(g)),str('='+str(alp[pt_col]+str(g))+'/'+str(alp[pt_col]+'$'+str(pt_st+1))),informat2_b)
         #       print(str('B='+str(alp[td_col]+str(g))+'/'+str(alp[td_col]+'$'+str(td_st+1))))
            elif g in [pt_st+3,pt_st+4,pt_st+5,pt_st+6,pt_st+8,pt_st+10,pt_st+12,pt_st+14]:
                worksheet.write_formula(str(alp[pt_col+1]+str(g)),str('='+str(alp[pt_col]+str(g))+'/'+str(alp[pt_col]+'$'+str(pt_st+1))),informat2_n)
         #       print(str('N='+str(alp[td_col]+str(g))+'/'+str(alp[td_col]+'$'+str(td_st+1))))
            else:
                gg=1
       
        # worksheet.conditional_format(str(alp[pt_col+1]+str(pt_st+11)), {'type': 'cell','criteria': '<','value': 0,'format': pt_cell_format_1})    
        # worksheet.conditional_format(str(alp[pt_col+1]+str(pt_st+11)), {'type': 'cell','criteria': '>','value': 0,'format': pt_cell_format_2})
         
         worksheet.conditional_format(str(alp[pt_col+1]+str(pt_st+11)), {'type':'cell','criteria': 'between','minimum': inv[f][4],
                                      'maximum': inv[f][5],'format': pt_cell_format_2})
        
         worksheet.conditional_format(str(alp[pt_col+1]+str(pt_st+11)), {'type':'cell','criteria': 'not between','minimum':  inv[f][4], 
                                           'maximum':  inv[f][5],
                                           'format':   pt_cell_format_1})
           
         worksheet.conditional_format(str(alp[pt_col+1]+str(pt_st+2)), {'type':'cell','criteria': 'between','minimum':  inv[f][6], 
                                           'maximum':  inv[f][7],
                                           'format':   pt_cell_format_2})
        
         worksheet.conditional_format(str(alp[pt_col+1]+str(pt_st+2)), {'type':'cell','criteria': 'not between','minimum':  inv[f][6], 
                                           'maximum':  inv[f][7],
                                           'format':   pt_cell_format_1})
         if automatic==True: 
             worksheet.merge_range(str(alp[pt_col]+str(pt_st+17)+":"+alp[pt_col+1]+str(pt_st+17)), "None" ,cell_format6_6)
         else:
             worksheet.merge_range(str(alp[pt_col]+str(pt_st+17)+":"+alp[pt_col+1]+str(pt_st+17)), "Trade EQTY" ,cell_format6_6)
         
         worksheet.data_validation(str(alp[pt_col]+str(pt_st+17)+":"+alp[pt_col+1]+str(pt_st+17)), {'validate': 'list',
                                            'source': ['Trade EQTY+FUT', 
                                                       'Trade EQTY', 
                                                       'Trade FUT',
                                                       'None'
                                                       ],
                                            'input_title': 'Enter a Trade Action',
                                            'input_message': 'Type'} )
         worksheet.conditional_format(str(alp[pt_col]+str(pt_st+17)+":"+alp[pt_col+1]+str(pt_st+17)),{'type': 'cell','criteria': '!=','value': '"N"','format': cell_format6_6})
         for tm in range(3):
             worksheet.write(str(alp[pt_col]+str(pt_st+20+tm)),c_st[0][f][tm],format1)
             worksheet.write_formula(str(alp[pt_col+1]+str(pt_st+20+tm)),str('='+str(alp[pt_col]+str(pt_st+20+tm))+'/'+str(alp[pt_col]+str(pt_st+1))),informat2_n)
         worksheet.conditional_format(str(alp[pt_col]+str(pt_st+22)), {'type':'formula',
                                                                      'criteria': str('=('+str(alp[pt_col]+str(pt_st+22))+'-'+str(alp[pt_col]+str(pt_st-1))+'+'+str(alp[pt_col]+str(11))+')<=0'),
                                                                      'format':   cell_format5_1})
    #     worksheet.conditional_format(str(alp[pt_col]+str(pt_st+17)),{'type': 'cell','criteria': 'equal to','value': '"N"','format': cell_format6_6})
        
         pt_col=pt_col+2
         
         
                           
    del [f ,g, gg, tm]
    
    worksheet.freeze_panes(18,2)
    #worksheet.write_comment('C10', 'Enter cash flow info below', {'start_col': 5,'start_row': 7, 'x_scale': 1.2, 'y_scale': 0.25, 'visible': True ,'font_size': 11, 'bold':True ,'color': '#FFCC99'})
    
    #pandaswb = writer.book
    #pandaswb.filename =  output_folder+'\\BatchCashCalc_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsm'
    #pandaswb.add_vba_project('C:/Program Files (x86)/WinPython/python-3.6.5.amd64/Scripts/vbaProject.bin')
    
    #pandaswb.save()        
    #pandaswb.close()                                                           
    worksheet.set_zoom(79)
    writer.save()
    workbook.close()
    
    
    mywb = px.load_workbook(output_folder1+'\\BatchCashCalc_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsm', keep_vba=True)
    mysheet = mywb.active
    
    border1=Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
      
    
    for j in range(4, 4+len(lst_fund)*2,2):
      #  print(j)
        mysheet.cell(row = 7, column = j).border = border1
        mysheet.cell(row = 16, column = j).border = border1
        mysheet.cell(row = 17, column = j).border = border1
           
    
    mysheet.protection.sheet = False
    mysheet.protection.password = 'Flower'
    
    mywb.save(output_folder1+'\\BatchCashCalc_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsm') 
        
    os.startfile(output_folder1)  
        
    
    #import win32com.client #import Dispatch
    
    #xl = win32com.client.Dispatch("Excel.Application")  # Set up excel
    #wb=xl.Workbooks.Open(Filename = output_folder+'BatchCashCalc_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsm')         # Open .xlsm file from step 2A
    #xl.Application.Run("Module1")                  # Run VBA_macro.bin
    #xl.Application.Run("auto_open")
    
    time_elapsed = datetime.now() - start_time 
    #print('Time elapsed (hh:mm:ss.ms) {}'.format(time_elapsed))
    msg_1="Batch cash calc \n generated!"
    return msg_1
    #workbook.add_vba_project('C:/IndexTrader/code/vbaProject.bin')
"""    
'******************************************************************************************************************************************************************************    
'                                                                   Create BPM Cash File
'******************************************************************************************************************************************************************************    
"""


def create_BPMcashfile(fnd_excp= ['DSALPC','OMCC01','OMCD01','OMCD02','OMCM01','OMCM02','PPSBTA','PPSBTB'], clear_cash=False,
                       user_file='C:\\IndexTrader\\required_inputs\\user_dictionary.csv',dir_out='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES',
                       ):
    
    import win32com.client #import Dispatch
    from write_excel import select_fund as sf
    import os
    import numpy as np
    import pandas as pd
    from datetime import datetime, timedelta
    from tkinter import filedialog
    from tkinter import Tk
    from tkinter import messagebox
    from write_excel import hedge_with_box as hf
    from write_excel import CreateBPMAttribute as bpm_attr
    import glob 
    import string

    lst_fund = sf(struc=False)   
    startDate = datetime.today()
   # startDate=startDate.replace(day=28)
   
    user_dict=pd.read_csv(user_file)
    dic_users=user_dict.set_index(['username']).T.to_dict('list')
   
    #fnd_excp= ['DSALPC','OMCC01','OMCD01','OMCD02','OMCM01','OMCM02','PPSBTA','PPSBTB']
    dirtoimport_file=dir_out
    folder_yr = datetime.strftime(startDate, "%Y")
    folder_mth = datetime.strftime(startDate, "%m")
    folder_day = datetime.strftime(startDate, "%d")

    output_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\BatchTrades\\')
   
    try:
        output_folder=str(output_folder+dic_users[os.environ.get("USERNAME").lower()][1])
    except:
        pass
    
    if clear_cash:
        root_file= max(glob.iglob(output_folder+'\\*.xlsm'), key=os.path.getmtime)
        root_file=root_file.replace('\\','/') 
    else:
        root = Tk()
        root.filename =  filedialog.askopenfilename(initialdir = output_folder,title = "Select Batch cash calc file",filetypes = (("jpeg files","*.xlsm"),("all files","*.*")))
        root_file=root.filename
        print (root.filename)
        root.withdraw()
        
    if root_file == '':
        msg1 = ['selected']
        msg=['no file']
        a_msg=['']
    else:
        
        sng = root_file
        sng_ = sng.split('/')
        sng_ = sng_[1:-1]
        output_folder = str('\\'+'\\'.join(sng_))
        print(output_folder)
     
        dirtooutput_file=str(output_folder+'\\CashFile\\')
        dirtooutput_fileF=str(output_folder+'\\FuturesFile\\')
        dirtooutput_fileBO=str(output_folder+'\\Batch Optimisation Reports\\')
        dirtooutput_fileET=str(output_folder+'\\Equity Trades\\')
        dirtooutput_fileAF=str(output_folder+'\\Attribute File\\')
        
        
        
        if not os.path.exists(dirtooutput_file):
            os.makedirs(dirtooutput_file)
        
        if not os.path.exists(dirtooutput_fileF):
            os.makedirs(dirtooutput_fileF)
       
        if not os.path.exists(dirtooutput_fileBO):
            os.makedirs(dirtooutput_fileBO)
       
        if not os.path.exists(dirtooutput_fileET):
            os.makedirs(dirtooutput_fileET)
        if not os.path.exists(dirtooutput_fileAF):
            os.makedirs(dirtooutput_fileAF)
            
        xl = win32com.client.Dispatch("Excel.Application")  # Set up excel
        wb=xl.Workbooks.Open(Filename = root_file)         # Open .xlsm file from step 2A
        ws=wb.Sheets[0]
        bsm=[]
        fsm=[]
       
        xl.Visible = False
        xl.ScreenUpdating = False
        xl.DisplayAlerts = False
        xl.EnableEvents = False
        fnd_chck=[]
        for h in range(3, len(lst_fund)*5+2, 2):
            #print(h)
            fnd_chck.append(ws.Cells(7, h).value)
        
        fnd_chck=[j for j in fnd_chck if j !=None]
        if  set(lst_fund)==set(fnd_chck):
        
        
            hg_with=[]
            for k in range(3, len(lst_fund)*2+2, 2):
                if ws.Cells(10, k).value=='Hedged Withdrawal':
                    print(k)
                    hg_with.append(ws.Cells(7, k).value)
            
           #pd.DataFrame(columns=['Port_code','Inflow_use','ActFlow'])
            if clear_cash:
                 chk_fut={}
                 for key in fnd_chck:
                     chk_fut[key]=0
            else:
                chk_fut=hf(hg_with,fnd_chck, snd=False)
     #      print(chk_fut)
            
            with open(str(dirtooutput_file+"BPM_Cash"+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'), "w",newline='\r\n') as fin:
            
                for j in range(3, len(fnd_chck)*2+2, 2):
            #        print(j)
                    fund = ws.Cells(7, j).value
                    bpm_cash=ws.Cells(80, j).value
                    if chk_fut[fund]==1:
                        bpm_futures=ws.Cells(28, j).value
                    else:
                        bpm_futures=ws.Cells(73, j).value
                    fut_code=ws.Cells(16, j).value
                    trd_typ=ws.Cells(81, j).value
                    if trd_typ in ['Trade EQTY' ,'Trade EQTY+FUT','Trade FUT']:
                      #  print(fut_code)
                        if fund in fnd_excp:
                            sf=fund+',ZAR,'+str(np.round(bpm_cash,15))+'\n'
                            sh=''.join(sf)
                        else:
                            sf=fund+',ZAR,'+str(np.round(bpm_cash,15))+'\n'+fund+','+fut_code+','+str(bpm_futures)+'\n'
                            sh=''.join(sf)
                        bsm=bsm+['Trade']    
                        fin.write(sh)
                    else:
                   #     print(fund)
                        msg="No cash file generated"
                        bsm=bsm+[msg]
            
            with open(str(dirtooutput_fileF+"Futures_"+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'), "w",newline='\r\n') as fut:
                
                for j in range(3, len(fnd_chck)*2+2, 2):
                    fund = ws.Cells(7, j).value
                    fut_code=ws.Cells(16, j).value
                    fut_no=np.abs(ws.Cells(58, j).value)
                    fut_value=ws.Cells(58, j+1).value
                    trd_typ=ws.Cells(81, j).value
                    if trd_typ in ['Trade FUT' ,'Trade EQTY+FUT']:
                        if fund in fnd_excp:
                            msg1='No futures file generated'
                            fsm=fsm+[msg1]
                        else:
                            sf=fut_code+','+fund+','+fut_value+','+str(fut_no)+',''Rebalance Portfolio,MP,,Trade at close\n'
                            sh=''.join(sf)
                            msg1="Futures file generated"
                            fsm=fsm+[msg1]
                            fut.write(sh)
                    else:
                       #print(fund)
                       msg1="No futures file generated"
                       fsm=fsm+[msg1]
                       #bsm=bsm+[msg]
            if clear_cash:
                a_msg=[''] 
            else:
                try:
                    a_msg=bpm_attr(ws,xl,startDate,dirtooutput_fileAF,dic_users,lst_fnds=fnd_chck)    
                except:
                    a_msg = ["Attribute file failed!"]
         
            wb.Close(False)
            xl.Visible=True
            xl.ScreenUpdating = True
            xl.DisplayAlerts = True
            xl.EnableEvents = True
            
            if all([elem == 'No cash file generated' for elem in bsm]):
                os.remove(str(dirtooutput_file+"BPM_Cash"+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'))
                msg=list([msg,str("Did you to enter a trade action?")])
            else:
                msg=["BPM Cash File created"]
                os.startfile(dirtooutput_file)  
              
            if all([elem == 'No futures file generated' for elem in fsm]):
                os.remove(str(dirtooutput_fileF+"Futures_"+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'))
                msg1=list([msg1])
            else:
                msg1=["Futures file generated"]
                #os.startfile(dirtooutput_file)  
        else:
            messagebox.showerror('Fund mismatch', 'Please note mismatch in funds selected.\nCheck that your Flows file matches the Batch cash calc file')
            msg=['Check']
            msg1=['Funds selected']
        try:
            a_msg=a_msg
        except:
            a_msg=['']
         
        
        del xl
    return '\r'.join(list(set(msg+msg1))+a_msg)
    

"""    
'******************************************************************************************************************************************************************************    
'******************************************************************************************************************************************************************************    
"""


"""    
'******************************************************************************************************************************************************************************    
'                                                                   Create Cash Flow Validity Fx 
'******************************************************************************************************************************************************************************    
"""



def cash_flow_validity_fx(cash_flows_eff,newest_cash,startDate,lst_fund, bf=0.005):

    import pandas as pd
    import numpy as np
        
#str(dirtoimport_file+newest)
#newest
#newest=str(dirtoimport_file+'UFMPosCash20190128.xls')
#startDate=startDate.replace(day=28)

#    cash_xls = pd.read_excel(newest_cash,sheet_name='Cash', 
#                             converters={'Settle Date': pd.to_datetime, 'Trade date':pd.to_datetime,
#                                                'Portfolio':str, 'Type':str, 
#                                                'Security name':str,
#                                                'Security Code':str,
#                                                'Quantity':float,
#                                                ' +/-':str,
#                                                'Amount': float},
#                            )

    cash_xls = pd.read_excel(newest_cash, sheet_name='Detail', 
                           converters={'Portfolio': str,
                                       'Instrument': str,
                                       'Currency': str,
                                       'Type':str,
                                       'Sign':str,
                                       'Qty':float,
                                       'Amt Current':float,
                                       'Amt Gross':float,
                                       'Trade date':pd.to_datetime,
                                       'Settle Date': pd.to_datetime},)
    
    cash_xls.columns = [col.strip()  for col in cash_xls.columns]
    cash_xls=cash_xls.drop(['Currency','Amt Gross'],axis=1)
    cash_xls.columns = ['Portfolio','Security Code','Type','+/-','Quantity','Amount','Trade date',	'Settle Date']

    
    cash_xlssub = (cash_xls.copy())[((cash_xls.Type.isin(['CSFLOW','CSHOUT','CSHINJ','CSHWTHD']))&(cash_xls.Portfolio.isin(lst_fund))
                                    &(pd.to_datetime(cash_xls['Settle Date'])==pd.to_datetime(startDate.date()))
                                    &(pd.to_datetime(cash_xls['Settle Date'])==pd.to_datetime(startDate.date())))]
    cash_xlssub.loc[:,'SysFlow']=np.where(cash_xlssub['+/-']=='-',-1.0*cash_xlssub.Amount.values, 
                                   np.where(cash_xlssub['+/-']=='+', cash_xlssub.Amount.values, cash_xlssub.Amount.values))
    
    cash_xlssub=cash_xlssub[['Trade date', 'Portfolio','Type', 'SysFlow']]
    cash_xlssub.columns=['Trade_date','Port_code','Type','SysFlow']
    #cash_xlssub_agg=cash_xlssub.groupby(['Port_code','Type']).agg({'SysFlow':'sum'})
    cash_xlssub_agg=cash_xlssub.groupby(['Port_code']).agg({'SysFlow':'sum'})
    # BL: added logic to group cash flows
    cash_xlssub_agg['Type']=np.where(cash_xlssub_agg['SysFlow']<0,'CSHWTHD', 
                                   np.where(cash_xlssub_agg['SysFlow']>0, 'CSHINJ', ''))
                            
           
    cash_flows_eff['Type']=np.where(cash_flows_eff['Inflow']<0,'CSHWTHD', 
                                   np.where(cash_flows_eff['Inflow']>0, 'CSHINJ', ''))
    
    cash_flows_eff_agg=cash_flows_eff.groupby(['Port_code','Type']).agg({'Inflow':'sum'})
    #BL: Added logic to group cash flows
    if len(cash_xlssub_agg)==0:
        cash_xlssub_agg=cash_xlssub_agg.drop(['Type'],axis=1)
    else:
        cash_xlssub_agg=cash_xlssub_agg.groupby(['Port_code','Type']).agg({'SysFlow':'sum'})
    
    csh_tab=pd.concat([cash_xlssub_agg,cash_flows_eff_agg], axis=1)
    csh_tab=csh_tab.reset_index().fillna(0)
    
    
    def cash_flw_flg(man_flow, sys_flow, buffer=bf,x=0):
        if man_flow==0:
            if sys_flow==0:
                return ['No flow',0,0][x]
            elif sys_flow!=0:
                return ['Valid flow, use system flow',sys_flow,0][x]
            else:
                return ['Weird',-999,0][x]
        elif man_flow!=0:
            if sys_flow==0:
                return ['Valid flow, use man flow',man_flow,man_flow][x]
            elif sys_flow!=0:
                if (abs(sys_flow/man_flow -1)<=buffer):
                    return ['Duplicated flow, use system',sys_flow,0][x]
                else:
                    return ['Valid flow, use man flow',man_flow,man_flow][x]
            else:
                return ['Weird',-999,0][x]
        else:
            return ['No flow',0,0,0][x]
                
        
    csh_tab['ValidCashFlag']=csh_tab.apply(lambda r: (cash_flw_flg(r.Inflow,r.SysFlow,bf,0)),axis=1)
    csh_tab['ActFlow']=csh_tab.apply(lambda r: (cash_flw_flg(r.Inflow,r.SysFlow,bf,1)),axis=1)       
    csh_tab['Inflow_use']= csh_tab.apply(lambda r: (cash_flw_flg(r.Inflow,r.SysFlow,bf,2)),axis=1)       
    
    csh_tab_agg=csh_tab.groupby(['Port_code']).agg({'Inflow_use':'sum','ActFlow':'sum'})
    csh_tab_agg=csh_tab_agg.reset_index()
    return[csh_tab,csh_tab_agg]
    
"""    
'******************************************************************************************************************************************************************************    
'******************************************************************************************************************************************************************************    
"""


"""
Created on Mon Feb 25 10:51:22 2019

@author: tmfelang2
"""
"""
' This function takes outputs from BPM (Batch optimisation report, cash file and trade list) and reformats to be compatible with
  post-opt recon requirements (IT)
  Drops re-formatted files into listener folder
  Also archives files into Trades_Archive folder
"""
def selt_fund():
    import tkinter 
    from tkinter import Label
    from tkinter import filedialog
    from tkinter import Entry
    from tkinter import Checkbutton
    from tkinter import IntVar
    from tkinter import Button
    from tkinter import mainloop
    #from tkinter import *
    coot = tkinter.Toplevel()
    coot.geometry("200x150+600+400")
    Label(coot, text="Please select files to drop:", font="Helvetica 10 bold").grid(row=0, sticky='w')
    cf = IntVar()
    Checkbutton(coot, text="Cash File", variable=cf,font="Helvetica 9").grid(row=1, sticky='w')
    bo = IntVar()
    Checkbutton(coot, text="Batch Optimisation Report", variable=bo,font="Helvetica 9").grid(row=2, sticky='w')
    trd = IntVar()
    Checkbutton(coot, text="Trade List", variable=trd,font="Helvetica 9").grid(row=3, sticky='w')

    Button(coot, text='OK', command=coot.quit, font="Helvetica 10").grid(row=4, pady=10, padx=80)
       
    
    coot.mainloop()
    
    coot.withdraw()
    print(cf.get())
    
    if (cf.get()==1):
       cx=True
    else:
       cx=False

    if (bo.get()==1):
        bx=True
    else:
        bx=False
    
    if (trd.get()==1):
        tx=True
    else:
        tx=False
    del [coot, cf, bo, trd]
    return [cx,bx,tx]
    

def BPM_output_loads():

    import pandas as pd
    
    import os
    import tkinter 
    from tkinter import Label
    from tkinter import filedialog
    from tkinter import Entry
    from tkinter import Checkbutton
    from tkinter import IntVar
    from tkinter import Button
    from tkinter import mainloop
    from write_excel import selt_fund as selt_fund 
    
    import datetime as dt
    from datetime import datetime, timedelta
    
    startDate = datetime.today()
      
    folder_yr = datetime.strftime(startDate, "%Y")
    folder_mth = datetime.strftime(startDate, "%m")
    folder_day = datetime.strftime(startDate, "%d")
    
    user_dict=pd.read_csv('C:\\IndexTrader\\required_inputs\\user_dictionary.csv')
    dic_users=user_dict.set_index(['username']).T.to_dict('list')
    
    IT_folder= '\\\\za.investment.int\\DFS\\SSApps\\FileTransfer\\BPMReconTool\\Listen'
  #  '\\\\za.investment.int\\DFS\\SSApps\\FileTransfer\\BPMReconTool\\Listen'
   
    def selectfiles(path, choose="Choose your Cash file"):
        
        root=tkinter.Tk()
        root.filename =  filedialog.askopenfilename(initialdir = path,title = choose)
        #print (root.filename)
        root.withdraw()
        
        print(choose[12:len(choose)])
        return root.filename
        
    
    #File directories
        
    xyz=selt_fund()
    msg=''
    
    if xyz[0]:
    
        dirtoimport_file='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES'
        cashfile_import_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\BatchTrades\\'+dic_users[os.environ.get("USERNAME").lower()][1]+'\\CashFile')
        cash_file=selectfiles(path=cashfile_import_folder)    
        if cash_file=='':
            xyz[0]=False
        else:
            bpm_cashfile=pd.read_csv(cash_file,index_col = None)
            bpm_cashfile.to_csv(str(IT_folder+'\\Cash Holdings\\BPM_Cash'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'),header = 1, index=False)
            
    else:
        pass
    
    if xyz[1]:
    
        dirtoimport_file='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES'
        batchopt_import_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\BatchTrades\\'+dic_users[os.environ.get("USERNAME").lower()][1]+'\\Batch Optimisation Reports')
        batchopt_file=selectfiles(choose="Choose your Batch Optimisation file",path=batchopt_import_folder) 
        if batchopt_file=='':
            xyz[1]=False
        else:
            batchoptimisationfile=pd.read_excel(batchopt_file, usecols='D:P', header = 1)
             
            #Edit Batch Optimisation report
            batchoptimisationfile = batchoptimisationfile.dropna(axis=0, subset=['Portfolio']) #if the portfolio column is empty then drop the row
            batchoptimisationfile['Active Risk'] *= 100  #multiply active risk column by 100
            batchoptimisationfile['Turnover'] *= 100  #multiply turover column by 100
           # batchoptimisationfile.to_csv('\\\\ZAWCAPP63\\OMIGSADataHub\\Workflows\\Listen\\ReconToolNewServer\\Batch Optimisation\\Batch Optimisation.csv', index=False)
            batchoptimisationfile.to_csv(str(IT_folder+'\\Batch Optimisation\\Batch Optimisation'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'), index=False)
         #  str('Batch_optimisation_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv')
    else: 
        pass
   
    if  xyz[2]:
    
        dirtoimport_file='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES'
        trdlist_import_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\BatchTrades\\'+dic_users[os.environ.get("USERNAME").lower()][1]+'\\Equity Trades')
        trd_file=selectfiles(choose="Choose your Trade file", path=trdlist_import_folder)
        if trd_file=='':
            xyz[2]=False
        else:
            tradelist=pd.read_csv(trd_file, header = 1)
            tradelist.to_csv(str(IT_folder+'\\Trades\\TradeList'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'), index=False)

    else:
        pass
    
    
    if ((xyz[0]==False)&(xyz[1]==False)&(xyz[2]==False)):
        msg='Please select \n file(s). \n No file(s) imported!'
        print(msg)
       
    else:
        
       #Drop files in Listening Folder
        #batchoptimisationfile.to_csv('\\\\ZAWCAPP63\\OMIGSADataHub\\Workflows\\Listen\\ReconTool\\Batch Optimisation\\Batch Optimisation.csv', index=False)
        #tradelist.to_csv('\\\\ZAWCAPP63\\OMIGSADataHub\\Workflows\\Listen\\ReconTool\\Trades\\TradeList.csv', index=False)
        #bpm_cashfile.to_csv('\\\\ZAWCAPP63\\OMIGSADataHub\\Workflows\\Listen\\ReconTool\\Cash Holdings\\BPM_Cash.csv', index=False)
        
    
        
        #archive the files imported
        dirtoexport_file='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES\\TRADES_ARCHIVE'
        
        if xyz[1]:
            batchopt_archive_folder=str('\\'.join([dirtoexport_file ,folder_yr, folder_mth,folder_day])+'\\BatchOptimisation')
            if not os.path.exists(batchopt_archive_folder):
                os.makedirs(batchopt_archive_folder)
            batchoptimisationfile.to_csv('\\'.join([batchopt_archive_folder,str('Batch_optimisation_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv')]), index=False)
            msg=msg+'Batch Opt,'
        
        if xyz[2]:
            trdlist_archive_folder=str('\\'.join([dirtoexport_file ,folder_yr, folder_mth,folder_day])+'\\TradeList')
            if not os.path.exists(trdlist_archive_folder):
                os.makedirs(trdlist_archive_folder)
            tradelist.to_csv('\\'.join([trdlist_archive_folder,str('Tradelist_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv')]), index=False)
            msg=msg+' Trade file, '
            
        if xyz[0]:
            cashfile_archive_folder=str('\\'.join([dirtoexport_file ,folder_yr, folder_mth,folder_day])+'\\CashFile')
            if not os.path.exists(cashfile_archive_folder):
                os.makedirs(cashfile_archive_folder)
            bpm_cashfile.to_csv('\\'.join([cashfile_archive_folder,str('Cashfile_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv')]), index=False)
            msg=msg+' Cash File '
        
        if (xyz[0]&xyz[1]&xyz[2]):
            msg='All files \n generated & saved'
        else:
            msg=msg+'imported!'
            
    return msg

"""
'******************************************************************************************************************************************************************************    
'******************************************************************************************************************************************************************************    
"""



"""
'******************************************************************************************************************************************************************************    
                                                                        Round Down Function
'******************************************************************************************************************************************************************************    
"""

def round_down(n, decimals=0):
    import math
    multiplier = 10 ** decimals
    return math.floor(n * multiplier) / multiplier


"""
'******************************************************************************************************************************************************************************    
                                                                        Hedged Withdrawal Check box function
'******************************************************************************************************************************************************************************    
"""


def hedge_with_box(chx_flw,lst_fnd, snd=True):
    import tkinter 
    from tkinter import Label
    from tkinter import filedialog
    from tkinter import Entry
    from tkinter import Checkbutton
    from tkinter import IntVar
    from tkinter import Button
    from tkinter import mainloop
    from tkinter import messagebox
    #from write_excel import select_fund as sf
    import numpy as np
    
    #lst_fnd= sf(struc=False) 
    
    if snd:
        hg_with_lst=(chx_flw[chx_flw.ActFlow<0].Port_code)[(chx_flw[chx_flw.ActFlow<0].Port_code).isin(lst_fnd)]
        hgt=np.int(len(hg_with_lst)/3*100)+50
    #from tkinter import *
    else:
        hg_with_lst=chx_flw
        hgt=np.int(len(hg_with_lst)/3*100)+60
    
    
        
    if len(hg_with_lst)>0:
        if snd:
            pass
        else:
            messagebox.showinfo("Temporary hedge withdrawal", "It appears as if you are temporarily hedging a withdrawal! \n Please select funds to export with current futures position!")

        coot = tkinter.Toplevel()
        coot.geometry(str("300x"+str(hgt)+"+600+450"))
        coot.title('Hedged Withdrawals')
        if snd:
            coot.title('Hedged Withdrawals')
            Label(coot, text="Please select funds with hedged withdrawals:", font="Helvetica 10 bold").grid(row=0, sticky='w')
        else:
            Label(coot, text="Export funds with current futures:", font="Helvetica 10 bold").grid(row=0, sticky='w')
                
        i=1
        var_categories = {}
     #   chckbx_categories = {}
        for fnd in hg_with_lst:
            cf = IntVar()
            Checkbutton(coot, text=fnd, variable=cf,font="Helvetica 9").grid(row=i, sticky='w')
            var_categories[fnd] = cf
            i+=1
            del cf
            
        Button(coot, text='OK', command=coot.quit, font="Helvetica 10").grid(row=i+1, pady=10, padx=80)
        del i
        #print(var_categories)   
        
        coot.mainloop()
        coot.withdraw()
    
        hdg_chck_lst={}
        
        for key in lst_fnd:
            hdg_chck_lst[key]=0
        
        for key,val in var_categories.items():   
   #         print(key)
   #         print(val.get())
            hdg_chck_lst[key]=val.get()
        
                
    else:
        hdg_chck_lst={}
        for key in lst_fnd:   
            hdg_chck_lst[key]=0
        
            
    return hdg_chck_lst

#hhgg=hedge_with_box(chx_flw,lst_fund)
        
"""
'******************************************************************************************************************************************************************************    
                                                                        Cash drop for clear cash fx
'******************************************************************************************************************************************************************************    
"""

def clear_cash_fx_drop(dir_imp='\\\\za.investment.int\\dfs\\dbshared\\DFM\\TRADES', lis_fld='\\\\za.investment.int\\DFS\\SSApps\\FileTransfer\\BPMReconTool\\Listen', user_file='C:\\IndexTrader\\required_inputs\\user_dictionary.csv'):    
    
    import datetime as dt
    from datetime import datetime, timedelta
    import pandas as pd
    import os
    import glob
    
    dirtoimport_file=dir_imp
    IT_folder= lis_fld
    startDate = datetime.today()
    folder_yr = datetime.strftime(startDate, "%Y")
    folder_mth = datetime.strftime(startDate, "%m")
    folder_day = datetime.strftime(startDate, "%d")
    user_dict=pd.read_csv(user_file)
    dic_users=user_dict.set_index(['username']).T.to_dict('list')
    
    cashfile_import_folder=str('\\'.join([dirtoimport_file ,folder_yr, folder_mth,folder_day])+'\\BatchTrades\\'+dic_users[os.environ.get("USERNAME").lower()][1]+'\\CashFile')
    cash_file=max(glob.iglob(cashfile_import_folder+'\\*.csv'), key=os.path.getmtime)
    if cash_file=='':
        msg='Error no file found'
    else:
        bpm_cashfile=pd.read_csv(cash_file,index_col = None)
        bpm_cashfile.to_csv(str(IT_folder+'\\Cash Holdings\\BPM_Cash'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.csv'),header = 1, index=False)
        msg='File drop to listner folder'
        os.startfile(str(IT_folder+'\\Cash Holdings\\')) 
    return msg

"""
'******************************************************************************************************************************************************************************    
                                                                        Cash Balances (t..t+3)
'******************************************************************************************************************************************************************************    
"""

def cash_forecast(dirtoimport_cashbal, startDate ,lst_fund):
   import pandas as pd
    
   import glob
   import os
   
   from bdateutil import isbday
   import bdateutil as bdut
   import holidays as hol      
   
   
   newest_cashbal=max(glob.iglob(dirtoimport_cashbal+'*.xls'), key=os.path.getmtime)
        #str(dirtoimport_file+newest)
        #newest
        
   bal_xls = pd.read_excel(newest_cashbal,sheet_name='Balances',converters={'Portfolio':str, 'Price Date': pd.to_datetime, 
        'Inst Type':str, 
        'Inst Name':str,
        'ISIN':str,
        'Instrument':str,
        'Quote Close':float,
        'Qty':float,
        'Market Val':float,
        'Delta':float,	
        'Origin':str},
        )
   bal_xls=bal_xls[bal_xls['Portfolio/Currency'].isin([str(f+'/ZAR')for f in lst_fund ])]
   bal_xls.loc[:,'Port_code']=bal_xls['Portfolio/Currency'].apply(lambda x: x.replace('/ZAR',''))
   bal_xls=bal_xls.drop(['Portfolio/Currency'], axis=1)                                 
   bal_xls=bal_xls[['Port_code','D','D + 1','D + 2','D + 3'] ]
   bal_xls=pd.pivot_table(bal_xls,  columns = 'Port_code')
   
   dtes=pd.bdate_range(start=startDate.date()+bdut.relativedelta(bdays=+1, holidays=hol.SouthAfrica()), end=startDate.date()+bdut.relativedelta(bdays=+3, holidays=hol.SouthAfrica()), holidays =hol.SouthAfrica())
            
   return [bal_xls, dtes]
    
    
    

    
def special_fund_load(fnd_lst,ans, fnd_dir='C:\\IndexTrader\\required_inputs\\', fnd_dict='fund_dictionary_global.csv'):
    import tkinter 
    from tkinter import Label
    from tkinter import Checkbutton
    from tkinter import IntVar
    from tkinter import Button
    from tkinter import Entry
    from tkinter import messagebox
    #from write_excel import select_fund as sf
    import numpy as np
    import pandas as pd
    
    if ans:
    
        fnd_csv=pd.read_csv(str(fnd_dir+fnd_dict))
        fnd_csv['TradingAccount']= np.where(fnd_csv.TradingAccount.isnull(),'None', fnd_csv.TradingAccount.values)
        dic_om_index=fnd_csv.set_index(['FundCode']).T.to_dict('list')
        hgt=np.int(len(fnd_lst)/3*100)+100
    
        tac_with_lst=fnd_lst
        if (all(g in fnd_csv['FundCode'].unique() for g in tac_with_lst)):
           
       
            coot = tkinter.Toplevel()
            coot.geometry(str("400x"+str(hgt)+"+600+450"))
            coot.title('Trading Accounts')
          #  Label(coot, text="", font="Helvetica 10 bold").grid(row=0, sticky="w")
            Label(coot, text="Please select fund(s) you require trading account(s) for:", font="Helvetica 10 bold").grid(row=0, columnspan=2,sticky="w")
            Label(coot, text="Fund", font="Helvetica 10 bold").grid(row=1, column=0,sticky="w")
            Label(coot, text="Trading Account", font="Helvetica 10 bold").grid(row=1, column=1,sticky="w")
            
                   
            i=1
            var_categories = {}
            trd_categories ={}
         #   chckbx_categories = {}
            if len(tac_with_lst)==1:
                cf = IntVar()
                Checkbutton(coot, text=tac_with_lst, variable=cf,font="Helvetica 9").grid(row=i+3,  column=0,sticky='w')
                entryText = tkinter.StringVar()
                e1=Entry(coot,font="Helvetica 9",textvariable=entryText).grid(row=i+3, column=1,sticky="w")
                entryText.set( dic_om_index[tac_with_lst[0]][4])
                var_categories[tac_with_lst[0]] = cf
                trd_categories[tac_with_lst[0]]= entryText
                print(entryText)
                del [cf,entryText]
            else:    
                for fnd in tac_with_lst:
                    cf = IntVar()
                    Checkbutton(coot, text=fnd, variable=cf,font="Helvetica 9").grid(row=i+3, sticky='w')
                    entryText = tkinter.StringVar()
                    e1=Entry(coot,font="Helvetica 9",textvariable=entryText).grid(row=i+3, column=1,sticky="w")
                    entryText.set(dic_om_index[fnd][4] )
                    var_categories[fnd] = cf
                    trd_categories[fnd] =entryText
                    print(entryText)
                    i+=1
                    del [cf,entryText]
            Label(coot, text= 'Trading Account(s):Enter alternative account code(s) if required', font="Helvetica 7" ,bg='pink').grid(row=i+5, column=0, columnspan = 2, sticky='w')   
            Button(coot, text='OK', command=coot.quit, font="Helvetica 10").place(relx=0.5, rely=0.88, anchor="c")
            del i
            #print(var_categories)   
            coot.mainloop()
            coot.withdraw()
    
            trd_chck_lst={}
            fnd_chck_lst={}
            
            for key in tac_with_lst:
                trd_chck_lst[key]=0
                fnd_chck_lst[key]=dic_om_index[key][4]
                
            for key,val in var_categories.items():   
                trd_chck_lst[key]=val.get()
            
            for key,val in trd_categories.items():   
                fnd_chck_lst[key]=val.get()
              
            for key in tac_with_lst:    
                if trd_chck_lst[key]==1:
                    if fnd_chck_lst[key] in ['None','']:
                        fnd_chck_lst[key]=key
                    else:
                        pass
                else:
                    fnd_chck_lst[key]=key
            msg='Yebo!!'
        else:
            msg='Fund selected is not part of Index Funds'
            fnd_chck_lst=[]
    else:
        msg='No trading accounts selected'
        fnd_chck_lst=[]
                        
    #print(msg)
    return [fnd_chck_lst,msg]
    
    
    

#special_fund_load(['OMRTMF','DSALPC','OMRTMF','OMSI01','MFEQTY'],ans=True, fnd_dir='C:\\IndexTrader\\required_inputs\\', fnd_dict='fund_dictionary_global.csv')

"""
'******************************************************************************************************************************************************************************    
                                                                        Create Attribute File (BPM)
'******************************************************************************************************************************************************************************    
"""

def CreateBPMAttribute(ws,xl,startDate,dirtooutput_fileAF,dic_users,lst_fnds):
    
    import pandas as pd
    import numpy as np
    import openpyxl
    import os
    import tkinter
    
    rowcount=ws.UsedRange.Rows.Count
    colcount=ws.UsedRange.Columns.Count
    lst_fnds=[]
    eq_trade=[]
    trd_act=[]
    for i in range(3,colcount,2): #start at column 3, step by 2 through all 'active' columns
     #   print(i)
        lst_fnds.append(xl.Cells(7,i).value) #reads row 7
        #lst_fnds = list(filter(None, lst_fnds)) # Filter out the None
        eq_trade.append(xl.Cells(63,i+1).value) #This will add 1 to the column 3 start
        trd_act.append(xl.Cells(81,i).value) # This will still start at column 3
    #wb.Close()
    
    keep=['Trade EQTY+FUT','Trade EQTY']
    trd_act_s=((pd.Series(trd_act)).isin(keep)).tolist()
    lst_fnds=(pd.Series(lst_fnds)[trd_act_s]).tolist()
    eq_trade=(pd.Series(eq_trade)[trd_act_s]).tolist()
    
    hgt=np.int(len(lst_fnds)/3*100)+70
        
    coot = tkinter.Toplevel()
    if len(lst_fnds)<10:
        coot.geometry(str("350x"+str(hgt)+"+550+500"))
    else:
        coot.geometry(str("350x"+str(hgt)+"+550+80"))
    coot.title('Trade type selection')
    tkinter.Label(coot, text="Please select trade action (BPM):", font="Helvetica 10 bold").grid(row=0,  sticky='w')
            
    i=1
    var_categories = {}
     #   chckbx_categories = {}
    for x in range(0,(len(lst_fnds))):
     #   print(x)
        cf = tkinter.IntVar()
        fnd=lst_fnds[x]
        tkinter.Label(coot, text=fnd, font="Helvetica 10 bold").grid(row=i,column=0, sticky='w',padx='10')
        tr_typ = tkinter.StringVar(coot)
    #    print(eq_trade[x].split(" ")[0])
        if eq_trade[x].split(" ")[0]=="":
            eq_trd_ac="ALLOWALL"
        else:
            eq_trd_ac=eq_trade[x].split(" ")[0]
            
        tr_typ.set(eq_trd_ac) # default value
        if tr_typ.get()=='BUY':
            var2="SELL"
            var3="ALLOWALL"
        elif tr_typ.get()=='SELL':
            var2="BUY"
            var3="ALLOWALL"
        else:
            var2="BUY"
            var3="SELL"
            
        
        w = tkinter.OptionMenu(coot, tr_typ, eq_trd_ac, var2, var3)
     #   w.config(bg = "white") 
        w.configure(background="white", activebackground="white")
        w["menu"].configure(bg="white")
        w.grid(row=i,column=1, sticky='w')
    
        var_categories[fnd] = tr_typ
        del cf
        i+=1
        
    if len(lst_fnds)>1:   
        tkinter.Button(coot, text='OK', command=coot.quit, font="Helvetica 10").grid(row=i+3,column=0,sticky='e',padx='20',pady='20')
    else:
        tkinter.Button(coot, text='OK', command=coot.quit, font="Helvetica 10").grid(row=i+3,column=0,sticky='e',padx='20',pady='10')
    del i
    #print(var_categories)   
    
    coot.mainloop()
    coot.withdraw()
    
    trd_type_lst={}   
    for key,val in var_categories.items():   
        trd_type_lst[key]=val.get()
            
    #Attribute File
  #  user_dict2=pd.read_csv('C:\\IndexTrader\\bpm_attribute\\user_dictionary_bpm.csv')
  #  dic_users2=user_dict2.set_index(['username']).T.to_dict('list')
    
    #OPENPYXL  
    wb = openpyxl.Workbook()
    sheet = wb.active
    c1 = sheet.cell(row = 1, column = 1)
    c1.value = '<ATTRIBUTE NAME>'
    c2 = sheet.cell(row = 2, column = 1)
    c2.value = '<OWNER>'
    c3 = sheet.cell(row = 3, column = 1)
    c3.value = '<ASSOCIATION>'
    c4 = sheet.cell(row = 4, column = 1)
    c4.value = '<TYPE>'
    c5 = sheet.cell(row = 5, column = 1)
    c5.value = '<MAXIMUM AGE>'
    c6 = sheet.cell(row = 6, column = 1)
    c6.value = '<PORTFOLIO>'
    c7 = sheet.cell(row = 1, column = 2)
    c7.value = (startDate.strftime('%Y%m%d-%H-%M-%S') +' ' + os.getlogin())
    c8 = sheet.cell(row = 2, column = 2)
    c8.value = (dic_users[os.environ.get("USERNAME").lower()][2])
    c9 = sheet.cell(row = 3, column = 2)
    c9.value = 'Miscellaneous'
    c10 = sheet.cell(row = 4, column = 2)
    c10.value = 'Text'
    c11 = sheet.cell(row = 5, column = 2)
    c11.value = 'None'
    c12 = sheet.cell(row = 6, column = 2)
    c12.value = '<PORTFOLIO OWNER>'
    c13 = sheet.cell(row = 6, column = 3)
    c13.value = '<VALUE>'
    
    for i in range(0,len(lst_fnds)):
        if i <= (len(lst_fnds)-1):
            b1=sheet.cell(row = 7+i, column = 1)
            b1.value = lst_fnds[i]
            b2=sheet.cell(row = 7+i, column = 3)
         #   b2.value = str(dic_users[os.environ.get("USERNAME").lower()][2] +'/'+ lst_fnds[i]+"_"+trd_type_lst[lst_fnds[i]])
            b2.value = str(dic_users['tmfelang2'][2] +'/'+ lst_fnds[i]+"_"+trd_type_lst[lst_fnds[i]])
            
            b3=sheet.cell(row = 7+i, column = 2)
            b3.value = 'ITAdmin'
        
    wb.save(('\\'.join([dirtooutput_fileAF,str('AttributeFile_'+startDate.strftime('%Y%m%d %H-%M-%S')+'_'+dic_users[os.environ.get("USERNAME").lower()][1]+'.xlsx')])))
    msg='Attribute File created'
    return [msg]